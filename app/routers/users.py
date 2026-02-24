# SPDX-License-Identifier: MIT
# Copyright (c) 2026 JAEHYUK CHO
import io
import uuid
from typing import Optional
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, status
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, field_validator
from sqlalchemy.orm import Session as DBSession
import openpyxl

from ..database import get_db
from ..models import User, UserExtraField, UserExtraValue, _now
from ..auth import get_current_user
from ..rbac import require_admin, can_manage_user, SUPER_ADMIN, ADMIN, USER
from ..crypto import crypto

router = APIRouter(prefix="/api/admin/users", tags=["users"])


class UserCreate(BaseModel):
    username: str
    password: str
    role: str = USER
    email: Optional[str] = None
    is_active: int = 1

    @field_validator("password")
    @classmethod
    def validate_password(cls, v):
        if len(v) < 8:
            raise ValueError("Password must be at least 8 characters")
        return v

    @field_validator("role")
    @classmethod
    def validate_role(cls, v):
        if v not in (SUPER_ADMIN, ADMIN, USER):
            raise ValueError("Invalid role")
        return v


class UserUpdate(BaseModel):
    username: Optional[str] = None
    password: Optional[str] = None
    role: Optional[str] = None
    email: Optional[str] = None
    is_active: Optional[int] = None


def _decrypt_email(email: Optional[str]) -> Optional[str]:
    """암호화된 이메일은 복호화, 평문(레거시)은 그대로 반환."""
    if not email:
        return email
    if crypto.is_encrypted(email):
        try:
            return crypto.decrypt(email)
        except Exception:
            return email
    return email


def _user_to_dict(user: User) -> dict:
    return {
        "id": user.id,
        "username": user.username,
        "email": _decrypt_email(user.email),
        "role": user.role,
        "is_active": user.is_active,
        "created_at": user.created_at,
        "updated_at": user.updated_at,
    }


@router.get("")
async def list_users(
    page: int = 1,
    page_size: int = 50,
    q: Optional[str] = None,
    current_user: User = Depends(require_admin),
    db: DBSession = Depends(get_db),
):
    query = db.query(User)
    if q:
        query = query.filter(User.username.contains(q))
    total = query.count()
    users = query.order_by(User.created_at.desc()).offset((page - 1) * page_size).limit(page_size).all()
    return {
        "data": [_user_to_dict(u) for u in users],
        "total": total,
        "page": page,
        "page_size": page_size,
    }


@router.post("", status_code=201)
async def create_user(
    body: UserCreate,
    current_user: User = Depends(require_admin),
    db: DBSession = Depends(get_db),
):
    # ADMIN은 SUPER_ADMIN 생성 불가
    if current_user.role == ADMIN and body.role == SUPER_ADMIN:
        raise HTTPException(status_code=403, detail="Cannot create SUPER_ADMIN")
    if db.query(User).filter(User.username == body.username).first():
        raise HTTPException(status_code=409, detail="Username already exists")
    user = User(
        id=str(uuid.uuid4()),
        username=body.username,
        email=crypto.encrypt(body.email) if body.email else None,
        password_hash=crypto.hash_password(body.password),
        role=body.role,
        is_active=body.is_active,
    )
    db.add(user)
    db.commit()
    db.refresh(user)
    return {"data": _user_to_dict(user), "message": "created"}


@router.get("/export-xlsx")
async def export_users_xlsx(
    current_user: User = Depends(require_admin),
    db: DBSession = Depends(get_db),
):
    users = db.query(User).order_by(User.created_at).all()
    fields = db.query(UserExtraField).order_by(UserExtraField.sort_order).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Users"

    headers = ["id", "username", "email", "role", "is_active"] + [f.field_name for f in fields]
    ws.append(headers)

    for u in users:
        vals_map = {v.field_id: v for v in u.extra_values}
        row = [u.id, u.username, _decrypt_email(u.email) or "", u.role, u.is_active]
        for f in fields:
            v = vals_map.get(f.id)
            if v and v.value:
                if f.is_sensitive:
                    try:
                        row.append(crypto.decrypt(v.value))
                    except Exception:
                        row.append("[DECRYPT_ERROR]")
                else:
                    row.append(v.value)
            else:
                row.append("")
        ws.append(row)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=users.xlsx"},
    )


@router.post("/import-xlsx")
async def import_users_xlsx(
    file: UploadFile = File(...),
    current_user: User = Depends(require_admin),
    db: DBSession = Depends(get_db),
):
    content = await file.read()
    if content[:4] != b"PK\x03\x04":
        raise HTTPException(status_code=400, detail="Invalid xlsx file")

    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="Empty file")

    headers = [str(h).strip() if h else "" for h in rows[0]]
    fields_by_name = {f.field_name: f for f in db.query(UserExtraField).all()}

    created, updated, skipped = 0, 0, 0
    for row in rows[1:]:
        row_dict = {headers[i]: row[i] for i in range(min(len(headers), len(row)))}
        username = str(row_dict.get("username", "") or "").strip()
        if not username:
            continue

        uid = str(row_dict.get("id", "") or "").strip()
        role = str(row_dict.get("role", USER) or USER).strip()
        if role not in (SUPER_ADMIN, ADMIN, USER):
            role = USER

        # ADMIN은 SUPER_ADMIN 행 건너뜀
        if current_user.role == ADMIN and role == SUPER_ADMIN:
            skipped += 1
            continue

        existing = (db.query(User).filter(User.id == uid).first() if uid else None) or \
                   db.query(User).filter(User.username == username).first()

        if existing:
            if not can_manage_user(current_user, existing):
                skipped += 1
                continue
            if row_dict.get("role"):
                existing.role = role
            if row_dict.get("email") is not None:
                _em = str(row_dict["email"]) if row_dict["email"] else None
                existing.email = crypto.encrypt(_em) if _em else None
            if row_dict.get("password"):
                existing.password_hash = crypto.hash_password(str(row_dict["password"]))
            if row_dict.get("is_active") is not None:
                existing.is_active = int(row_dict["is_active"])
            existing.updated_at = _now()
            updated += 1
        else:
            pw = str(row_dict.get("password", "") or "").strip()
            if not pw:
                pw = "changeme!"
            _em = str(row_dict.get("email", "") or "") or None
            user = User(
                id=uid or str(uuid.uuid4()),
                username=username,
                email=crypto.encrypt(_em) if _em else None,
                password_hash=crypto.hash_password(pw),
                role=role,
                is_active=int(row_dict.get("is_active", 1) or 1),
            )
            db.add(user)
            created += 1

        # extra fields
        u_obj = existing or user
        for fname, field in fields_by_name.items():
            if fname in row_dict and row_dict[fname] is not None:
                val_str = str(row_dict[fname])
                if field.is_sensitive:
                    val_str = crypto.encrypt(val_str)
                uev = db.query(UserExtraValue).filter(
                    UserExtraValue.user_id == u_obj.id,
                    UserExtraValue.field_id == field.id,
                ).first()
                if uev:
                    uev.value = val_str
                else:
                    db.add(UserExtraValue(
                        id=str(uuid.uuid4()),
                        user_id=u_obj.id,
                        field_id=field.id,
                        value=val_str,
                    ))

    db.commit()
    return {"message": "imported", "created": created, "updated": updated, "skipped": skipped}


@router.get("/{user_id}")
async def get_user(
    user_id: str,
    current_user: User = Depends(require_admin),
    db: DBSession = Depends(get_db),
):
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    return {"data": _user_to_dict(user)}


@router.patch("/{user_id}")
async def update_user(
    user_id: str,
    body: UserUpdate,
    current_user: User = Depends(require_admin),
    db: DBSession = Depends(get_db),
):
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    if not can_manage_user(current_user, user):
        raise HTTPException(status_code=403, detail="Cannot manage this user")

    if body.username is not None:
        if db.query(User).filter(User.username == body.username, User.id != user_id).first():
            raise HTTPException(status_code=409, detail="Username already exists")
        user.username = body.username
    if body.password is not None:
        if len(body.password) < 8:
            raise HTTPException(status_code=400, detail="Password too short")
        user.password_hash = crypto.hash_password(body.password)
    if body.role is not None:
        if current_user.role == ADMIN and body.role == SUPER_ADMIN:
            raise HTTPException(status_code=403, detail="Cannot set SUPER_ADMIN role")
        user.role = body.role
    if body.email is not None:
        user.email = crypto.encrypt(body.email) if body.email else None
    if body.is_active is not None:
        user.is_active = body.is_active
    user.updated_at = _now()
    db.commit()
    return {"data": _user_to_dict(user), "message": "updated"}


@router.delete("/{user_id}", status_code=204)
async def delete_user(
    user_id: str,
    current_user: User = Depends(require_admin),
    db: DBSession = Depends(get_db),
):
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    if user.id == current_user.id:
        raise HTTPException(status_code=400, detail="Cannot delete yourself")
    if not can_manage_user(current_user, user):
        raise HTTPException(status_code=403, detail="Cannot manage this user")
    db.delete(user)
    db.commit()


@router.get("/{user_id}/field-values")
async def get_field_values(
    user_id: str,
    current_user: User = Depends(require_admin),
    db: DBSession = Depends(get_db),
):
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    fields = db.query(UserExtraField).order_by(UserExtraField.sort_order).all()
    vals_map = {v.field_id: v for v in user.extra_values}
    result = []
    for f in fields:
        v = vals_map.get(f.id)
        raw = v.value if v else None
        if raw and f.is_sensitive:
            try:
                raw = crypto.decrypt(raw)
            except Exception:
                raw = "[DECRYPT_ERROR]"
        result.append({
            "field_id": f.id,
            "field_name": f.field_name,
            "label": f.label,
            "field_type": f.field_type,
            "is_sensitive": f.is_sensitive,
            "value": raw,
        })
    return {"data": result}


@router.patch("/{user_id}/field-values")
async def update_field_values(
    user_id: str,
    body: dict,
    current_user: User = Depends(require_admin),
    db: DBSession = Depends(get_db),
):
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    if not can_manage_user(current_user, user):
        raise HTTPException(status_code=403, detail="Cannot manage this user")

    for field_id, value in body.items():
        field = db.query(UserExtraField).filter(UserExtraField.id == field_id).first()
        if not field:
            continue
        val_str = str(value) if value is not None else None
        if val_str and field.is_sensitive:
            val_str = crypto.encrypt(val_str)
        uev = db.query(UserExtraValue).filter(
            UserExtraValue.user_id == user_id,
            UserExtraValue.field_id == field_id,
        ).first()
        if uev:
            uev.value = val_str
        else:
            db.add(UserExtraValue(
                id=str(uuid.uuid4()),
                user_id=user_id,
                field_id=field_id,
                value=val_str,
            ))
    db.commit()
    return {"message": "updated"}
