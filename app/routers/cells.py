# SPDX-License-Identifier: MIT
# Copyright (c) 2026 JAEHYUK CHO
import uuid
import json
from typing import Optional
from fastapi import APIRouter, Depends, HTTPException, Request
from pydantic import BaseModel
from sqlalchemy.orm import Session as DBSession

from ..database import get_db
from ..models import (
    Workspace, WorkspaceSheet, WorkspaceCell, TemplateSheet, TemplateColumn,
    ChangeLog, _now
)
from ..auth import get_current_user
from ..rbac import require_user, is_admin_or_above
from ..models import User
from ..ws_hub import hub
from .templates import _style_to_css, _range_to_jss, _freeze_to_cols, _freeze_to_rows, _pt_to_px
from openpyxl.utils import get_column_letter

router = APIRouter(prefix="/api/workspaces", tags=["cells"])

MAX_ROWS = 10000
MAX_COLS = 16384


class PatchItem(BaseModel):
    row: int
    col: int
    value: Optional[str] = None
    style: Optional[str] = None  # style JSON 문자열 (선택)
    comment: Optional[str] = None  # 셀 메모 (선택)
    hyperlink: Optional[str] = None  # URL 하이퍼링크 (선택)


class PatchRequest(BaseModel):
    patches: list[PatchItem]


async def _apply_patches(
    workspace_id: str,
    sheet_id: str,
    patches: list[PatchItem],
    current_user: User,
    db: DBSession,
    broadcast: bool = True,
) -> list:
    ws = db.query(Workspace).filter(Workspace.id == workspace_id).first()
    if not ws:
        raise HTTPException(status_code=404, detail="Workspace not found")

    ws_sheet = db.query(WorkspaceSheet).filter(
        WorkspaceSheet.id == sheet_id,
        WorkspaceSheet.workspace_id == workspace_id,
    ).first()
    if not ws_sheet:
        raise HTTPException(status_code=404, detail="Sheet not found")

    # CLOSED 체크
    if ws.status == "CLOSED" and not is_admin_or_above(current_user):
        raise HTTPException(status_code=403, detail="Workspace is closed")

    # 컬럼 readonly 맵
    tmpl_sheet = db.query(TemplateSheet).filter(TemplateSheet.id == ws_sheet.template_sheet_id).first()
    readonly_cols: set[int] = set()
    if tmpl_sheet:
        for c in tmpl_sheet.columns:
            if c.is_readonly:
                readonly_cols.add(c.col_index)

    # 시트 보호 상태 확인
    sheet_protected = bool(ws_sheet.sheet_protected)

    applied = []
    for p in patches:
        if not (0 <= p.row < MAX_ROWS):
            continue
        if not (0 <= p.col < MAX_COLS):
            continue
        # readonly 컬럼: 일반 사용자 거부
        if p.col in readonly_cols and not is_admin_or_above(current_user):
            continue
        # 시트 보호 + 셀 잠금: 일반 사용자 거부
        if sheet_protected and not is_admin_or_above(current_user):
            cell_check = db.query(WorkspaceCell).filter(
                WorkspaceCell.sheet_id == sheet_id,
                WorkspaceCell.row_index == p.row,
                WorkspaceCell.col_index == p.col,
            ).first()
            if cell_check and cell_check.style:
                try:
                    s = json.loads(cell_check.style)
                    if s.get("locked"):
                        continue
                except Exception:
                    pass

        existing = db.query(WorkspaceCell).filter(
            WorkspaceCell.sheet_id == sheet_id,
            WorkspaceCell.row_index == p.row,
            WorkspaceCell.col_index == p.col,
        ).first()
        old_value = existing.value if existing else None

        # 하이퍼링크 보안 검증
        safe_hyperlink = None
        if p.hyperlink is not None:
            hl = p.hyperlink.strip()
            if hl and not any(hl.lower().startswith(proto) for proto in ('http://', 'https://', 'mailto:', '#')):
                hl = None  # 허용되지 않은 프로토콜 차단
            safe_hyperlink = hl if hl else None

        if existing:
            if p.value is not None:
                existing.value = p.value
            if p.style is not None:
                existing.style = p.style
            if p.comment is not None:
                existing.comment = p.comment if p.comment else None
            if p.hyperlink is not None:
                existing.hyperlink = safe_hyperlink
            existing.updated_by = current_user.id
            existing.updated_at = _now()
        else:
            db.add(WorkspaceCell(
                id=str(uuid.uuid4()),
                sheet_id=sheet_id,
                row_index=p.row,
                col_index=p.col,
                value=p.value,
                style=p.style,
                comment=p.comment if p.comment else None,
                hyperlink=safe_hyperlink,
                updated_by=current_user.id,
                updated_at=_now(),
            ))

        # 변경 이력 (값 변경만)
        if p.value is not None and old_value != p.value:
            db.add(ChangeLog(
                id=str(uuid.uuid4()),
                workspace_id=workspace_id,
                sheet_id=sheet_id,
                user_id=current_user.id,
                row_index=p.row,
                col_index=p.col,
                old_value=old_value,
                new_value=p.value,
            ))
        applied.append({"row": p.row, "col": p.col, "value": p.value, "style": p.style, "comment": p.comment, "hyperlink": p.hyperlink})

    db.commit()
    return applied


@router.get("/{workspace_id}/sheets/{sheet_id}/snapshot")
async def get_snapshot(
    workspace_id: str,
    sheet_id: str,
    current_user: User = Depends(require_user),
    db: DBSession = Depends(get_db),
):
    ws = db.query(Workspace).filter(Workspace.id == workspace_id).first()
    if not ws:
        raise HTTPException(status_code=404, detail="Workspace not found")
    ws_sheet = db.query(WorkspaceSheet).filter(
        WorkspaceSheet.id == sheet_id,
        WorkspaceSheet.workspace_id == workspace_id,
    ).first()
    if not ws_sheet:
        raise HTTPException(status_code=404, detail="Sheet not found")

    cells = db.query(WorkspaceCell).filter(WorkspaceCell.sheet_id == sheet_id).all()

    # 컬럼 메타
    tmpl_sheet = db.query(TemplateSheet).filter(TemplateSheet.id == ws_sheet.template_sheet_id).first()
    num_cols = len(tmpl_sheet.columns) if tmpl_sheet else 0
    # 셀 데이터 + 병합 데이터에서 컬럼 수 보정 (열 삽입으로 확장된 경우 포함)
    if cells:
        max_cell_col = max((c.col_index for c in cells), default=-1) + 1
        num_cols = max(num_cols, max_cell_col)
    if ws_sheet.merges:
        try:
            from openpyxl.utils import range_boundaries
            for rng in json.loads(ws_sheet.merges):
                _, _, max_col, _ = range_boundaries(rng)
                num_cols = max(num_cols, max_col)
        except Exception:
            pass
    num_cols = max(num_cols, 5)  # 최소 5열
    max_row = max((c.row_index for c in cells), default=-1) + 1
    num_rows = max(max_row, 100)

    # 2D 배열 + 스타일 맵 + 메모 맵 + 숫자 서식 맵 + 하이퍼링크 맵
    grid = [[""] * num_cols for _ in range(num_rows)]
    styles: dict = {}
    comments: dict = {}
    num_formats: dict = {}
    hyperlinks: dict = {}
    for c in cells:
        if c.row_index < num_rows and c.col_index < num_cols:
            grid[c.row_index][c.col_index] = c.value or ""
            if c.style:
                try:
                    s = json.loads(c.style)
                    if s:
                        cell_name = f"{get_column_letter(c.col_index + 1)}{c.row_index + 1}"
                        styles[cell_name] = _style_to_css(s)
                        if s.get('numFmt'):
                            num_formats[cell_name] = s['numFmt']
                except Exception:
                    pass
            if c.comment:
                cell_name = f"{get_column_letter(c.col_index + 1)}{c.row_index + 1}"
                comments[cell_name] = c.comment
            if c.hyperlink:
                cell_name = f"{get_column_letter(c.col_index + 1)}{c.row_index + 1}"
                hyperlinks[cell_name] = c.hyperlink

    # 병합 셀: xlsx range strings → jspreadsheet format
    merges: dict = {}
    if ws_sheet.merges:
        try:
            for rng in json.loads(ws_sheet.merges):
                result = _range_to_jss(rng)
                if result:
                    cell_name, dims = result
                    merges[cell_name] = dims
        except Exception:
            pass

    # 행 높이: pt → px
    row_heights_px: dict = {}
    if ws_sheet.row_heights:
        try:
            for k, v in json.loads(ws_sheet.row_heights).items():
                row_heights_px[k] = _pt_to_px(float(v))
        except Exception:
            pass

    # 열 너비: col_widths (px)
    col_widths_px: dict = {}
    if ws_sheet.col_widths:
        try:
            col_widths_px = json.loads(ws_sheet.col_widths)
        except Exception:
            pass

    # 틀 고정
    freeze_columns = _freeze_to_cols(ws_sheet.freeze_panes)
    freeze_rows = _freeze_to_rows(ws_sheet.freeze_panes)

    # 조건부 서식
    conditional_formats = []
    if ws_sheet.conditional_formats:
        try:
            conditional_formats = json.loads(ws_sheet.conditional_formats)
        except Exception:
            pass

    # 데이터 유효성 검사
    data_validations = []
    if ws_sheet.data_validations:
        try:
            data_validations = json.loads(ws_sheet.data_validations)
        except Exception:
            pass

    # 숨겨진 행/열
    hidden_rows_list = []
    if ws_sheet.hidden_rows:
        try:
            hidden_rows_list = json.loads(ws_sheet.hidden_rows)
        except Exception:
            pass
    hidden_cols_list = []
    if ws_sheet.hidden_cols:
        try:
            hidden_cols_list = json.loads(ws_sheet.hidden_cols)
        except Exception:
            pass

    return {
        "data": {
            "cells": grid,
            "num_rows": num_rows,
            "num_cols": num_cols,
            "merges": merges,
            "row_heights": row_heights_px,
            "col_widths": col_widths_px,
            "freeze_columns": freeze_columns,
            "freeze_rows": freeze_rows,
            "styles": styles,
            "comments": comments,
            "num_formats": num_formats,
            "conditional_formats": conditional_formats,
            "hyperlinks": hyperlinks,
            "hidden_rows": hidden_rows_list,
            "hidden_cols": hidden_cols_list,
            "data_validations": data_validations,
            "sheet_protected": bool(ws_sheet.sheet_protected),
            "print_settings": json.loads(ws_sheet.print_settings) if ws_sheet.print_settings else None,
            "outline_rows": json.loads(ws_sheet.outline_rows) if ws_sheet.outline_rows else {},
            "outline_cols": json.loads(ws_sheet.outline_cols) if ws_sheet.outline_cols else {},
        }
    }


@router.post("/{workspace_id}/sheets/{sheet_id}/patches")
async def http_patches(
    workspace_id: str,
    sheet_id: str,
    body: PatchRequest,
    current_user: User = Depends(require_user),
    db: DBSession = Depends(get_db),
):
    applied = await _apply_patches(workspace_id, sheet_id, body.patches, current_user, db)
    if applied:
        import asyncio
        msg = {
            "type": "batch_patch",
            "sheet_id": sheet_id,
            "patches": applied,
            "updated_by": current_user.username,
        }
        asyncio.create_task(hub.broadcast(workspace_id, msg))
    return {"message": "ok", "applied": len(applied)}


# ── Row insert / delete ──────────────────────────────────────

class RowInsertRequest(BaseModel):
    row_index: int
    count: int = 1
    direction: str = "above"  # "above" or "below"


class RowDeleteRequest(BaseModel):
    row_indices: list[int]


@router.post("/{workspace_id}/sheets/{sheet_id}/rows/insert")
async def insert_rows(
    workspace_id: str,
    sheet_id: str,
    body: RowInsertRequest,
    request: Request,
    current_user: User = Depends(require_user),
    db: DBSession = Depends(get_db),
):
    ws = db.query(Workspace).filter(Workspace.id == workspace_id).first()
    if not ws:
        raise HTTPException(status_code=404, detail="Workspace not found")
    if ws.status == "CLOSED" and not is_admin_or_above(current_user):
        raise HTTPException(status_code=403, detail="Workspace is closed")

    ws_sheet = db.query(WorkspaceSheet).filter(
        WorkspaceSheet.id == sheet_id,
        WorkspaceSheet.workspace_id == workspace_id,
    ).first()
    if not ws_sheet:
        raise HTTPException(status_code=404, detail="Sheet not found")

    insert_at = body.row_index
    count = max(1, min(body.count, 100))  # limit to 100 rows at a time

    # Shift existing cells down
    # Use negative offset first to avoid UNIQUE constraint conflicts,
    # then shift to final position
    db.execute(
        WorkspaceCell.__table__.update()
        .where(WorkspaceCell.sheet_id == sheet_id)
        .where(WorkspaceCell.row_index >= insert_at)
        .values(row_index=WorkspaceCell.row_index - 100000)
    )
    db.execute(
        WorkspaceCell.__table__.update()
        .where(WorkspaceCell.sheet_id == sheet_id)
        .where(WorkspaceCell.row_index < -90000)
        .values(row_index=WorkspaceCell.row_index + 100000 + count)
    )

    # Update merges JSON
    if ws_sheet.merges:
        try:
            merges = json.loads(ws_sheet.merges)
            updated = _shift_merges(merges, insert_at, count)
            ws_sheet.merges = json.dumps(updated)
        except Exception:
            pass

    # Update row_heights JSON
    if ws_sheet.row_heights:
        try:
            rh = json.loads(ws_sheet.row_heights)
            new_rh = {}
            for k, v in rh.items():
                ri = int(k)
                if ri >= insert_at:
                    new_rh[str(ri + count)] = v
                else:
                    new_rh[k] = v
            ws_sheet.row_heights = json.dumps(new_rh)
        except Exception:
            pass

    db.commit()

    # Broadcast to other clients
    import asyncio
    asyncio.create_task(hub.broadcast(workspace_id, {
        "type": "row_insert",
        "sheet_id": sheet_id,
        "row_index": insert_at,
        "count": count,
        "updated_by": current_user.username,
        "tab_id": request.headers.get("X-Tab-ID", ""),
    }, exclude=None))

    return {"message": "rows inserted", "row_index": insert_at, "count": count}


@router.post("/{workspace_id}/sheets/{sheet_id}/rows/delete")
async def delete_rows(
    workspace_id: str,
    sheet_id: str,
    body: RowDeleteRequest,
    request: Request,
    current_user: User = Depends(require_user),
    db: DBSession = Depends(get_db),
):
    ws = db.query(Workspace).filter(Workspace.id == workspace_id).first()
    if not ws:
        raise HTTPException(status_code=404, detail="Workspace not found")
    if ws.status == "CLOSED" and not is_admin_or_above(current_user):
        raise HTTPException(status_code=403, detail="Workspace is closed")

    ws_sheet = db.query(WorkspaceSheet).filter(
        WorkspaceSheet.id == sheet_id,
        WorkspaceSheet.workspace_id == workspace_id,
    ).first()
    if not ws_sheet:
        raise HTTPException(status_code=404, detail="Sheet not found")

    indices = sorted(set(body.row_indices))
    if not indices:
        return {"message": "no rows to delete"}

    # Delete cells in target rows
    for ri in indices:
        db.query(WorkspaceCell).filter(
            WorkspaceCell.sheet_id == sheet_id,
            WorkspaceCell.row_index == ri,
        ).delete()

    # Shift remaining cells up using raw SQL (two-step to avoid UNIQUE conflicts)
    # Step 1: move all cells above first deleted row to negative space
    first_deleted = indices[0]
    db.execute(
        WorkspaceCell.__table__.update()
        .where(WorkspaceCell.sheet_id == sheet_id)
        .where(WorkspaceCell.row_index > first_deleted)
        .values(row_index=WorkspaceCell.row_index - 100000)
    )
    # Step 2: compute correct new positions and move back
    # Each cell's new row = old_row - (number of deleted rows below it)
    from sqlalchemy import text
    for cell in db.query(WorkspaceCell).filter(
        WorkspaceCell.sheet_id == sheet_id,
        WorkspaceCell.row_index < -90000,
    ).all():
        original_row = cell.row_index + 100000
        offset = sum(1 for d in indices if d < original_row)
        cell.row_index = original_row - offset

    # Update merges JSON
    if ws_sheet.merges:
        try:
            merges = json.loads(ws_sheet.merges)
            for ri in reversed(indices):
                merges = _shift_merges(merges, ri, -1)
            ws_sheet.merges = json.dumps(merges)
        except Exception:
            pass

    # Update row_heights JSON
    if ws_sheet.row_heights:
        try:
            rh = json.loads(ws_sheet.row_heights)
            new_rh = {}
            for k, v in rh.items():
                ri = int(k)
                if ri in indices:
                    continue  # skip deleted rows
                offset = sum(1 for d in indices if d < ri)
                new_rh[str(ri - offset)] = v
            ws_sheet.row_heights = json.dumps(new_rh)
        except Exception:
            pass

    db.commit()

    import asyncio
    asyncio.create_task(hub.broadcast(workspace_id, {
        "type": "row_delete",
        "sheet_id": sheet_id,
        "row_indices": indices,
        "updated_by": current_user.username,
        "tab_id": request.headers.get("X-Tab-ID", ""),
    }, exclude=None))

    return {"message": "rows deleted", "count": len(indices)}


# ── Column insert / delete ────────────────────────────────────

class ColInsertRequest(BaseModel):
    col_index: int
    count: int = 1
    direction: str = "before"  # "before" or "after"


class ColDeleteRequest(BaseModel):
    col_indices: list[int]


@router.post("/{workspace_id}/sheets/{sheet_id}/cols/insert")
async def insert_cols(
    workspace_id: str,
    sheet_id: str,
    body: ColInsertRequest,
    request: Request,
    current_user: User = Depends(require_user),
    db: DBSession = Depends(get_db),
):
    ws = db.query(Workspace).filter(Workspace.id == workspace_id).first()
    if not ws:
        raise HTTPException(status_code=404, detail="Workspace not found")
    if ws.status == "CLOSED" and not is_admin_or_above(current_user):
        raise HTTPException(status_code=403, detail="Workspace is closed")

    ws_sheet = db.query(WorkspaceSheet).filter(
        WorkspaceSheet.id == sheet_id,
        WorkspaceSheet.workspace_id == workspace_id,
    ).first()
    if not ws_sheet:
        raise HTTPException(status_code=404, detail="Sheet not found")

    insert_at = body.col_index if body.direction == "before" else body.col_index + 1
    count = max(1, min(body.count, 50))

    # Shift existing cells right using two-step negative offset
    db.execute(
        WorkspaceCell.__table__.update()
        .where(WorkspaceCell.sheet_id == sheet_id)
        .where(WorkspaceCell.col_index >= insert_at)
        .values(col_index=WorkspaceCell.col_index - 100000)
    )
    db.execute(
        WorkspaceCell.__table__.update()
        .where(WorkspaceCell.sheet_id == sheet_id)
        .where(WorkspaceCell.col_index < -90000)
        .values(col_index=WorkspaceCell.col_index + 100000 + count)
    )

    # Update merges JSON
    if ws_sheet.merges:
        try:
            merges = json.loads(ws_sheet.merges)
            updated = _shift_merges_col(merges, insert_at, count)
            ws_sheet.merges = json.dumps(updated)
        except Exception:
            pass

    db.commit()

    import asyncio
    asyncio.create_task(hub.broadcast(workspace_id, {
        "type": "col_insert",
        "sheet_id": sheet_id,
        "col_index": insert_at,
        "count": count,
        "updated_by": current_user.username,
        "tab_id": request.headers.get("X-Tab-ID", ""),
    }, exclude=None))

    return {"message": "columns inserted", "col_index": insert_at, "count": count}


@router.post("/{workspace_id}/sheets/{sheet_id}/cols/delete")
async def delete_cols(
    workspace_id: str,
    sheet_id: str,
    body: ColDeleteRequest,
    request: Request,
    current_user: User = Depends(require_user),
    db: DBSession = Depends(get_db),
):
    ws = db.query(Workspace).filter(Workspace.id == workspace_id).first()
    if not ws:
        raise HTTPException(status_code=404, detail="Workspace not found")
    if ws.status == "CLOSED" and not is_admin_or_above(current_user):
        raise HTTPException(status_code=403, detail="Workspace is closed")

    ws_sheet = db.query(WorkspaceSheet).filter(
        WorkspaceSheet.id == sheet_id,
        WorkspaceSheet.workspace_id == workspace_id,
    ).first()
    if not ws_sheet:
        raise HTTPException(status_code=404, detail="Sheet not found")

    indices = sorted(set(body.col_indices))
    if not indices:
        return {"message": "no columns to delete"}

    # Delete cells in target columns
    for ci in indices:
        db.query(WorkspaceCell).filter(
            WorkspaceCell.sheet_id == sheet_id,
            WorkspaceCell.col_index == ci,
        ).delete()

    # Shift remaining cells left using two-step negative offset
    first_deleted = indices[0]
    db.execute(
        WorkspaceCell.__table__.update()
        .where(WorkspaceCell.sheet_id == sheet_id)
        .where(WorkspaceCell.col_index > first_deleted)
        .values(col_index=WorkspaceCell.col_index - 100000)
    )
    for cell in db.query(WorkspaceCell).filter(
        WorkspaceCell.sheet_id == sheet_id,
        WorkspaceCell.col_index < -90000,
    ).all():
        original_col = cell.col_index + 100000
        offset = sum(1 for d in indices if d < original_col)
        cell.col_index = original_col - offset

    # Update merges JSON
    if ws_sheet.merges:
        try:
            merges = json.loads(ws_sheet.merges)
            for ci in reversed(indices):
                merges = _shift_merges_col(merges, ci, -1)
            ws_sheet.merges = json.dumps(merges)
        except Exception:
            pass

    db.commit()

    import asyncio
    asyncio.create_task(hub.broadcast(workspace_id, {
        "type": "col_delete",
        "sheet_id": sheet_id,
        "col_indices": indices,
        "updated_by": current_user.username,
        "tab_id": request.headers.get("X-Tab-ID", ""),
    }, exclude=None))

    return {"message": "columns deleted", "count": len(indices)}


def _shift_merges_col(merges: list[str], at_col: int, shift: int) -> list[str]:
    """Shift merge ranges when columns are inserted/deleted."""
    from openpyxl.utils import range_boundaries
    updated = []
    for rng in merges:
        try:
            min_col, min_row, max_col, max_row = range_boundaries(rng)
            if shift > 0:
                # Insert: at_col is 0-based, range_boundaries returns 1-based
                at_col_1 = at_col + 1
                if min_col >= at_col_1:
                    min_col += shift
                if max_col >= at_col_1:
                    max_col += shift
            else:
                # Delete: at_col is 0-based, range_boundaries returns 1-based
                deleted_col_1 = at_col + 1
                if min_col <= deleted_col_1 <= max_col:
                    if max_col - min_col == 0:
                        continue  # entire merge deleted
                    max_col -= 1
                elif deleted_col_1 < min_col:
                    min_col -= 1
                    max_col -= 1
            start = f"{get_column_letter(min_col)}{min_row}"
            end = f"{get_column_letter(max_col)}{max_row}"
            if start != end:
                updated.append(f"{start}:{end}")
        except Exception:
            updated.append(rng)
    return updated


def _shift_merges(merges: list[str], at_row: int, shift: int) -> list[str]:
    """Shift merge ranges when rows are inserted/deleted."""
    from openpyxl.utils import range_boundaries
    updated = []
    for rng in merges:
        try:
            min_col, min_row, max_col, max_row = range_boundaries(rng)
            if shift > 0:
                # Insert: at_row is 0-based, min_row/max_row are 1-based
                # Row at 1-based N corresponds to 0-based index N-1
                # Shift rows whose 0-based index >= at_row, i.e., 1-based > at_row
                if min_row > at_row:
                    min_row += shift
                if max_row > at_row:
                    max_row += shift
            else:
                # Delete: remove ranges that overlap deleted row, shift others
                deleted_row = at_row + 1  # 1-based
                if min_row <= deleted_row <= max_row:
                    if max_row - min_row == 0:
                        continue  # entire merge deleted
                    max_row -= 1
                elif deleted_row < min_row:
                    min_row -= 1
                    max_row -= 1
            start = f"{get_column_letter(min_col)}{min_row}"
            end = f"{get_column_letter(max_col)}{max_row}"
            if start != end:
                updated.append(f"{start}:{end}")
        except Exception:
            updated.append(rng)
    return updated


class SortRequest(BaseModel):
    col_index: int
    ascending: bool = True


@router.post("/{workspace_id}/sheets/{sheet_id}/sort")
async def sort_sheet(
    workspace_id: str,
    sheet_id: str,
    body: SortRequest,
    request: Request,
    current_user: User = Depends(require_user),
    db: DBSession = Depends(get_db),
):
    ws = db.query(Workspace).filter(Workspace.id == workspace_id).first()
    if not ws:
        raise HTTPException(status_code=404, detail="Workspace not found")
    if ws.status == "CLOSED" and not is_admin_or_above(current_user):
        raise HTTPException(status_code=403, detail="Workspace is closed")

    ws_sheet = db.query(WorkspaceSheet).filter(
        WorkspaceSheet.id == sheet_id,
        WorkspaceSheet.workspace_id == workspace_id,
    ).first()
    if not ws_sheet:
        raise HTTPException(status_code=404, detail="Sheet not found")

    # 병합 셀 감지 - 행 간 병합이 있으면 정렬 거부
    if ws_sheet.merges:
        try:
            from openpyxl.utils import range_boundaries
            for rng in json.loads(ws_sheet.merges):
                _, min_row, _, max_row = range_boundaries(rng)
                if max_row > min_row:
                    raise HTTPException(status_code=400, detail="행 간 병합 셀이 있어 정렬할 수 없습니다")
        except HTTPException:
            raise
        except Exception:
            pass

    cells = db.query(WorkspaceCell).filter(WorkspaceCell.sheet_id == sheet_id).all()
    if not cells:
        return {"message": "no data to sort"}

    max_row = max(c.row_index for c in cells)
    row_values = {}
    for c in cells:
        if c.col_index == body.col_index:
            row_values[c.row_index] = c.value or ""

    row_indices = list(range(max_row + 1))

    def sort_key(ri):
        val = row_values.get(ri, "")
        if not val:
            return (1, "")
        try:
            return (0, float(val))
        except (ValueError, TypeError):
            return (0, val.lower())

    row_indices.sort(key=sort_key, reverse=not body.ascending)
    old_to_new = {old_ri: new_ri for new_ri, old_ri in enumerate(row_indices)}

    OFFSET = 100000
    for c in cells:
        c.row_index = old_to_new.get(c.row_index, c.row_index) + OFFSET
    db.flush()
    for c in cells:
        c.row_index -= OFFSET
    db.flush()

    if ws_sheet.row_heights:
        try:
            old_heights = json.loads(ws_sheet.row_heights)
            new_heights = {}
            for old_str, val in old_heights.items():
                new_ri = old_to_new.get(int(old_str), int(old_str))
                new_heights[str(new_ri)] = val
            ws_sheet.row_heights = json.dumps(new_heights)
        except Exception:
            pass

    db.commit()

    import asyncio
    asyncio.create_task(hub.broadcast(workspace_id, {
        "type": "reload", "updated_by": current_user.username,
    }, exclude=None))
    return {"message": "sorted"}
