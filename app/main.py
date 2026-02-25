# SPDX-License-Identifier: MIT
# Copyright (c) 2026 JAEHYUK CHO
import logging
import time
from pathlib import Path

from fastapi import FastAPI, Request, Depends
from fastapi.responses import HTMLResponse, RedirectResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from starlette.middleware.base import BaseHTTPMiddleware
from starlette.middleware.trustedhost import TrustedHostMiddleware

from .config import settings
from .database import get_db, db_session
from .auth import get_current_user, get_current_user_optional
from .models import User, Workspace, Template, ChangeLog
from .rbac import is_admin_or_above
from .routers import auth, users, user_fields, templates, workspaces, cells, websocket

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

BASE_DIR = Path(__file__).parent.parent
TEMPLATES_DIR = BASE_DIR / "web" / "templates"
STATIC_DIR = BASE_DIR / "web" / "static"

app = FastAPI(title="OpenSpace", docs_url="/api/docs", redoc_url=None)

# Static files
app.mount("/static", StaticFiles(directory=str(STATIC_DIR)), name="static")

# Templates
jinja = Jinja2Templates(directory=str(TEMPLATES_DIR))


# ------------------------------------------------------------------
# Security headers middleware
# ------------------------------------------------------------------
class SecurityHeadersMiddleware(BaseHTTPMiddleware):
    async def dispatch(self, request: Request, call_next):
        response = await call_next(request)
        response.headers["X-Content-Type-Options"] = "nosniff"
        response.headers["X-Frame-Options"] = "DENY"
        response.headers["X-XSS-Protection"] = "1; mode=block"
        response.headers["Referrer-Policy"] = "strict-origin-when-cross-origin"
        if not settings.debug:
            response.headers["Strict-Transport-Security"] = "max-age=31536000"
        csp = (
            "default-src 'self'; "
            # 'unsafe-eval' 필수: jspreadsheet 수식 엔진이 new Function()을 사용
            "script-src 'self' 'unsafe-inline' 'unsafe-eval'; "
            "style-src 'self' 'unsafe-inline'; "
            "font-src 'self'; "
            "img-src 'self' data:; "
            "connect-src 'self' ws: wss:;"
        )
        response.headers["Content-Security-Policy"] = csp
        return response


# ------------------------------------------------------------------
# CSRF middleware
# ------------------------------------------------------------------
CSRF_EXEMPT_PATHS = {"/api/auth/login", "/api/auth/logout"}
CSRF_SAFE_METHODS = {"GET", "HEAD", "OPTIONS"}


class CSRFMiddleware(BaseHTTPMiddleware):
    async def dispatch(self, request: Request, call_next):
        if (
            request.method not in CSRF_SAFE_METHODS
            and request.url.path not in CSRF_EXEMPT_PATHS
            and request.url.path.startswith("/api/")
        ):
            cookie_token = request.cookies.get("csrf_token")
            header_token = request.headers.get("X-CSRF-Token")
            if not cookie_token or cookie_token != header_token:
                from fastapi.responses import JSONResponse
                return JSONResponse(
                    status_code=403,
                    content={"detail": "CSRF token mismatch"},
                )
        return await call_next(request)


app.add_middleware(SecurityHeadersMiddleware)
app.add_middleware(CSRFMiddleware)
if "*" not in settings.allowed_hosts:
    app.add_middleware(TrustedHostMiddleware, allowed_hosts=settings.allowed_hosts)

# ------------------------------------------------------------------
# Routers
# ------------------------------------------------------------------
app.include_router(auth.router)
app.include_router(users.router)
app.include_router(user_fields.router)
app.include_router(templates.router)
app.include_router(workspaces.router)
app.include_router(cells.router)
app.include_router(websocket.router)


# ------------------------------------------------------------------
# Page routes (SSR)
# ------------------------------------------------------------------

@app.get("/", include_in_schema=False)
async def root(request: Request):
    with db_session() as db:
        user = get_current_user_optional(request, db)
    if user:
        return RedirectResponse("/dashboard")
    return RedirectResponse("/login")


@app.get("/login", response_class=HTMLResponse, include_in_schema=False)
async def login_page(request: Request):
    return jinja.TemplateResponse("login.html", {"request": request})


@app.get("/dashboard", response_class=HTMLResponse, include_in_schema=False)
async def dashboard(request: Request):
    with db_session() as db:
        user = get_current_user_optional(request, db)
        if not user:
            return RedirectResponse("/login")
        ws_count = db.query(Workspace).count()
        tmpl_count = db.query(Template).count()
        user_count = db.query(User).count()
        open_workspaces = db.query(Workspace).filter(Workspace.status == "OPEN").all()

        # 최근 변경 이력 (ADMIN+ 전용)
        recent_logs = []
        if is_admin_or_above(user):
            rows = (
                db.query(ChangeLog, User, Workspace)
                .join(User, ChangeLog.user_id == User.id)
                .join(Workspace, ChangeLog.workspace_id == Workspace.id)
                .order_by(ChangeLog.changed_at.desc())
                .limit(20)
                .all()
            )
            from openpyxl.utils import get_column_letter
            for log, log_user, log_ws in rows:
                recent_logs.append({
                    "username": log_user.username,
                    "workspace_name": log_ws.name,
                    "workspace_id": log_ws.id,
                    "cell": f"{get_column_letter(log.col_index + 1)}{log.row_index + 1}",
                    "old_value": log.old_value,
                    "new_value": log.new_value,
                    "changed_at": log.changed_at,
                })

        return jinja.TemplateResponse("dashboard.html", {
            "request": request,
            "current_user": user,
            "ws_count": ws_count,
            "tmpl_count": tmpl_count,
            "user_count": user_count,
            "open_workspaces": open_workspaces,
            "recent_logs": recent_logs,
            "is_admin": is_admin_or_above(user),
        })


@app.get("/admin/users", response_class=HTMLResponse, include_in_schema=False)
async def users_page(request: Request):
    with db_session() as db:
        user = get_current_user_optional(request, db)
        if not user:
            return RedirectResponse("/login")
        if not is_admin_or_above(user):
            return RedirectResponse("/dashboard")
        return jinja.TemplateResponse("users.html", {"request": request, "current_user": user})


@app.get("/admin/user-fields", response_class=HTMLResponse, include_in_schema=False)
async def user_fields_page(request: Request):
    with db_session() as db:
        user = get_current_user_optional(request, db)
        if not user:
            return RedirectResponse("/login")
        if not is_admin_or_above(user):
            return RedirectResponse("/dashboard")
        return jinja.TemplateResponse("user_fields.html", {"request": request, "current_user": user})


@app.get("/admin/templates", response_class=HTMLResponse, include_in_schema=False)
async def templates_page(request: Request):
    with db_session() as db:
        user = get_current_user_optional(request, db)
        if not user:
            return RedirectResponse("/login")
        if not is_admin_or_above(user):
            return RedirectResponse("/dashboard")
        return jinja.TemplateResponse("templates.html", {"request": request, "current_user": user})


@app.get("/admin/templates/{template_id}/edit", response_class=HTMLResponse, include_in_schema=False)
async def template_edit_page(template_id: str, request: Request):
    import json as _json
    from .models import Template as Tmpl
    with db_session() as db:
        user = get_current_user_optional(request, db)
        if not user:
            return RedirectResponse("/login")
        if not is_admin_or_above(user):
            return RedirectResponse("/dashboard")
        t = db.query(Tmpl).filter(Tmpl.id == template_id).first()
        if not t:
            return RedirectResponse("/admin/templates")
        sheets_data = []
        for s in sorted(t.sheets, key=lambda x: x.sheet_index):
            cols = []
            for c in sorted(s.columns, key=lambda x: x.col_index):
                cols.append({
                    "id": c.id,
                    "col_index": c.col_index,
                    "col_header": c.col_header,
                    "col_type": c.col_type,
                    "is_readonly": bool(c.is_readonly),
                    "width": c.width or 120,
                })
            sheets_data.append({"id": s.id, "sheet_index": s.sheet_index, "sheet_name": s.sheet_name, "columns": cols})
        return jinja.TemplateResponse("template_edit.html", {
            "request": request,
            "current_user": user,
            "template": t,
            "template_data_json": _json.dumps({"id": t.id, "name": t.name, "sheets": sheets_data}),
            "is_admin": is_admin_or_above(user),
        })


@app.get("/admin/workspaces", response_class=HTMLResponse, include_in_schema=False)
async def admin_workspaces_page(request: Request):
    with db_session() as db:
        user = get_current_user_optional(request, db)
        if not user:
            return RedirectResponse("/login")
        if not is_admin_or_above(user):
            return RedirectResponse("/dashboard")
        return jinja.TemplateResponse("workspaces.html", {"request": request, "current_user": user})


@app.get("/workspaces", response_class=HTMLResponse, include_in_schema=False)
async def workspace_list_page(request: Request):
    with db_session() as db:
        user = get_current_user_optional(request, db)
        if not user:
            return RedirectResponse("/login")
        workspaces_list = db.query(Workspace).order_by(Workspace.created_at.desc()).all()
        return jinja.TemplateResponse("workspace_list.html", {
            "request": request,
            "current_user": user,
            "workspaces": workspaces_list,
            "is_admin": is_admin_or_above(user),
        })


@app.get("/workspaces/{workspace_id}", response_class=HTMLResponse, include_in_schema=False)
async def workspace_page(workspace_id: str, request: Request):
    import json
    from .models import TemplateSheet, WorkspaceCell
    from openpyxl.utils import get_column_letter
    with db_session() as db:
        user = get_current_user_optional(request, db)
        if not user:
            return RedirectResponse("/login")
        ws = db.query(Workspace).filter(Workspace.id == workspace_id).first()
        if not ws:
            return RedirectResponse("/workspaces")

        # 시트 + 컬럼 메타 구성 (N+1 방지: 일괄 로드)
        ws_sheets = sorted(ws.sheets, key=lambda x: x.sheet_index)
        tmpl_sheet_ids = [s.template_sheet_id for s in ws_sheets if s.template_sheet_id]
        tmpl_sheets_map: dict = {}
        if tmpl_sheet_ids:
            for ts in db.query(TemplateSheet).filter(TemplateSheet.id.in_(tmpl_sheet_ids)).all():
                tmpl_sheets_map[ts.id] = ts

        sheets_data = []
        for s in ws_sheets:
            tmpl_sheet = tmpl_sheets_map.get(s.template_sheet_id)
            cols = []
            if tmpl_sheet:
                for c in sorted(tmpl_sheet.columns, key=lambda x: x.col_index):
                    cols.append({
                        "id": c.id,
                        "col_index": c.col_index,
                        "col_header": c.col_header,
                        "col_type": c.col_type,
                        "is_readonly": bool(c.is_readonly),
                        "width": c.width or 120,
                    })
            # 셀 데이터 + 병합에서 컬럼 수 보정 (열 삽입으로 확장된 경우 포함)
            from sqlalchemy import func
            max_col = db.query(func.max(WorkspaceCell.col_index)).filter(
                WorkspaceCell.sheet_id == s.id
            ).scalar()
            num_cols_from_cells = (max_col + 1) if max_col is not None else 0
            if s.merges:
                try:
                    import json as _j
                    from openpyxl.utils import range_boundaries
                    for rng in _j.loads(s.merges):
                        _, _, mc, _ = range_boundaries(rng)
                        num_cols_from_cells = max(num_cols_from_cells, mc)
                except Exception:
                    pass
            needed = max(num_cols_from_cells, 5)
            if not cols:
                for ci in range(needed):
                    cols.append({
                        "id": f"auto-{ci}",
                        "col_index": ci,
                        "col_header": get_column_letter(ci + 1),
                        "col_type": "text",
                        "is_readonly": False,
                        "width": 120,
                    })
            elif needed > len(cols):
                for ci in range(len(cols), needed):
                    cols.append({
                        "id": f"auto-{ci}",
                        "col_index": ci,
                        "col_header": get_column_letter(ci + 1),
                        "col_type": "text",
                        "is_readonly": False,
                        "width": 120,
                    })
            sheets_data.append({
                "id": s.id,
                "sheet_index": s.sheet_index,
                "sheet_name": s.sheet_name,
                "columns": cols,
            })

        workspace_data = {
            "id": ws.id,
            "name": ws.name,
            "status": ws.status,
            "sheets": sheets_data,
        }

        return jinja.TemplateResponse("workspace.html", {
            "request": request,
            "current_user": user,
            "workspace": ws,
            "workspace_data_json": json.dumps(workspace_data),
            "is_admin": is_admin_or_above(user),
        })
