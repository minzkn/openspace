# SPDX-License-Identifier: MIT
# Copyright (c) 2026 JAEHYUK CHO
import io
import json
import uuid
from typing import Optional
from fastapi import APIRouter, Depends, HTTPException, Request, UploadFile, File
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy.orm import Session as DBSession
import openpyxl
from openpyxl.utils import get_column_letter

from ..database import get_db
from ..models import (
    Template, TemplateSheet, TemplateColumn, TemplateCell,
    Workspace, WorkspaceSheet, WorkspaceCell, _now
)
from ..auth import get_current_user
from ..rbac import require_admin, require_user, ADMIN, SUPER_ADMIN
from ..models import User
from ..ws_hub import hub
from .templates import (
    _extract_cell_style, _apply_cell_style, _style_to_css,
    _range_to_jss, _freeze_to_cols, _pt_to_px,
    _get_theme_colors, _stringify_value, _parse_value_for_excel,
    _sanitize_xlsx,
)

router = APIRouter(tags=["workspaces"])


class WorkspaceCreate(BaseModel):
    name: str
    template_id: str


class WSSheetCreate(BaseModel):
    sheet_name: str = "새 시트"


class WSSheetUpdate(BaseModel):
    sheet_name: Optional[str] = None


class WorkspaceUpdate(BaseModel):
    name: Optional[str] = None


class MergesUpdate(BaseModel):
    merges: list[str]


def _ws_summary(ws: Workspace) -> dict:
    return {
        "id": ws.id,
        "name": ws.name,
        "template_id": ws.template_id,
        "status": ws.status,
        "created_by": ws.created_by,
        "closed_by": ws.closed_by,
        "closed_at": ws.closed_at,
        "created_at": ws.created_at,
        "updated_at": ws.updated_at,
        "sheet_count": len(ws.sheets),
    }


def _ws_detail(ws: Workspace, db: DBSession) -> dict:
    tmpl_sheet_ids = [s.template_sheet_id for s in ws.sheets if s.template_sheet_id]
    tmpl_sheets: dict[str, TemplateSheet] = {}
    if tmpl_sheet_ids:
        for ts in db.query(TemplateSheet).filter(TemplateSheet.id.in_(tmpl_sheet_ids)).all():
            tmpl_sheets[ts.id] = ts

    sheets = []
    for s in sorted(ws.sheets, key=lambda x: x.sheet_index):
        tmpl_sheet = tmpl_sheets.get(s.template_sheet_id)
        cols = []
        if tmpl_sheet:
            for c in sorted(tmpl_sheet.columns, key=lambda x: x.col_index):
                cols.append({
                    "id": c.id,
                    "col_index": c.col_index,
                    "col_header": c.col_header,
                    "col_type": c.col_type,
                    "col_options": json.loads(c.col_options) if c.col_options else None,
                    "is_readonly": c.is_readonly,
                    "width": c.width,
                })
        sheets.append({
            "id": s.id,
            "sheet_index": s.sheet_index,
            "sheet_name": s.sheet_name,
            "template_sheet_id": s.template_sheet_id,
            "columns": cols,
        })
    return {**_ws_summary(ws), "sheets": sheets}


# ---------------------------------------------------------------
# Admin routes
# ---------------------------------------------------------------

admin_router = APIRouter(prefix="/api/admin/workspaces")


@admin_router.get("")
async def admin_list_workspaces(
    current_user: User = Depends(require_admin),
    db: DBSession = Depends(get_db),
):
    workspaces = db.query(Workspace).order_by(Workspace.created_at.desc()).all()
    return {"data": [_ws_summary(w) for w in workspaces]}


@admin_router.post("", status_code=201)
async def create_workspace(
    body: WorkspaceCreate,
    current_user: User = Depends(require_admin),
    db: DBSession = Depends(get_db),
):
    tmpl = db.query(Template).filter(Template.id == body.template_id).first()
    if not tmpl:
        raise HTTPException(status_code=404, detail="Template not found")

    ws = Workspace(
        id=str(uuid.uuid4()),
        name=body.name,
        template_id=body.template_id,
        created_by=current_user.id,
    )
    db.add(ws)

    for sheet in tmpl.sheets:
        ws_sheet = WorkspaceSheet(
            id=str(uuid.uuid4()),
            workspace_id=ws.id,
            template_sheet_id=sheet.id,
            sheet_index=sheet.sheet_index,
            sheet_name=sheet.sheet_name,
            # 템플릿 시트 메타 복사
            merges=sheet.merges,
            row_heights=sheet.row_heights,
            col_widths=sheet.col_widths,
            freeze_panes=sheet.freeze_panes,
            conditional_formats=sheet.conditional_formats,
        )
        db.add(ws_sheet)
        # 템플릿 셀 복사 (formula 우선, style/comment 포함)
        for cell in sheet.cells:
            db.add(WorkspaceCell(
                id=str(uuid.uuid4()),
                sheet_id=ws_sheet.id,
                row_index=cell.row_index,
                col_index=cell.col_index,
                value=cell.formula if cell.formula else cell.value,
                style=cell.style,
                comment=cell.comment,
            ))

    db.commit()
    db.refresh(ws)
    return {"data": _ws_summary(ws), "message": "created"}


@admin_router.get("/{workspace_id}")
async def admin_get_workspace(
    workspace_id: str,
    current_user: User = Depends(require_admin),
    db: DBSession = Depends(get_db),
):
    ws = db.query(Workspace).filter(Workspace.id == workspace_id).first()
    if not ws:
        raise HTTPException(status_code=404, detail="Workspace not found")
    return {"data": _ws_detail(ws, db)}


@admin_router.patch("/{workspace_id}")
async def update_workspace(
    workspace_id: str,
    body: WorkspaceUpdate,
    current_user: User = Depends(require_admin),
    db: DBSession = Depends(get_db),
):
    ws = db.query(Workspace).filter(Workspace.id == workspace_id).first()
    if not ws:
        raise HTTPException(status_code=404, detail="Workspace not found")
    if body.name is not None:
        ws.name = body.name
    ws.updated_at = _now()
    db.commit()
    return {"data": _ws_summary(ws), "message": "updated"}


@admin_router.delete("/{workspace_id}", status_code=204)
async def delete_workspace(
    workspace_id: str,
    current_user: User = Depends(require_admin),
    db: DBSession = Depends(get_db),
):
    ws = db.query(Workspace).filter(Workspace.id == workspace_id).first()
    if not ws:
        raise HTTPException(status_code=404, detail="Workspace not found")
    db.delete(ws)
    db.commit()


class BatchDeleteRequest(BaseModel):
    ids: list[str]


@admin_router.post("/batch-delete")
async def batch_delete_workspaces(
    body: BatchDeleteRequest,
    current_user: User = Depends(require_admin),
    db: DBSession = Depends(get_db),
):
    targets = db.query(Workspace).filter(Workspace.id.in_(body.ids)).all()
    for w in targets:
        db.delete(w)
    db.commit()
    return {"deleted": len(targets)}


@admin_router.post("/{workspace_id}/close")
async def close_workspace(
    workspace_id: str,
    current_user: User = Depends(require_admin),
    db: DBSession = Depends(get_db),
):
    ws = db.query(Workspace).filter(Workspace.id == workspace_id).first()
    if not ws:
        raise HTTPException(status_code=404, detail="Workspace not found")
    if ws.status == "CLOSED":
        raise HTTPException(status_code=400, detail="Already closed")
    ws.status = "CLOSED"
    ws.closed_by = current_user.id
    ws.closed_at = _now()
    ws.updated_at = _now()
    db.commit()
    import asyncio
    asyncio.create_task(hub.broadcast(workspace_id, {"type": "workspace_status", "status": "CLOSED"}))
    return {"message": "closed"}


@admin_router.post("/{workspace_id}/reopen")
async def reopen_workspace(
    workspace_id: str,
    current_user: User = Depends(require_admin),
    db: DBSession = Depends(get_db),
):
    ws = db.query(Workspace).filter(Workspace.id == workspace_id).first()
    if not ws:
        raise HTTPException(status_code=404, detail="Workspace not found")
    ws.status = "OPEN"
    ws.closed_by = None
    ws.closed_at = None
    ws.updated_at = _now()
    db.commit()
    import asyncio
    asyncio.create_task(hub.broadcast(workspace_id, {"type": "workspace_status", "status": "OPEN"}))
    return {"message": "reopened"}


@admin_router.patch("/{workspace_id}/sheets/{sheet_id}/merges")
async def update_ws_sheet_merges(
    workspace_id: str,
    sheet_id: str,
    body: MergesUpdate,
    request: Request,
    current_user: User = Depends(require_admin),
    db: DBSession = Depends(get_db),
):
    """워크스페이스 시트 병합 저장"""
    ws_sheet = db.query(WorkspaceSheet).filter(
        WorkspaceSheet.id == sheet_id,
        WorkspaceSheet.workspace_id == workspace_id,
    ).first()
    if not ws_sheet:
        raise HTTPException(status_code=404, detail="Sheet not found")
    ws_sheet.merges = json.dumps(body.merges, ensure_ascii=False)
    db.commit()
    import asyncio
    asyncio.create_task(hub.broadcast(workspace_id, {
        "type": "sheet_config_updated",
        "sheet_id": sheet_id,
        "updated_by": current_user.username,
        "tab_id": request.headers.get("X-Tab-ID", ""),
    }, exclude=None))
    return {"message": "merges saved", "count": len(body.merges)}


class RowHeightsUpdate(BaseModel):
    row_heights: Optional[dict] = None  # e.g. {"0": 30, "5": 50} (row_index_str → pt)


class FreezeUpdate(BaseModel):
    freeze_panes: Optional[str] = None  # e.g. "C1" or null


class ConditionalFormatsUpdate(BaseModel):
    conditional_formats: Optional[str] = None  # JSON string of rules array


@admin_router.patch("/{workspace_id}/sheets/{sheet_id}/row-heights")
async def update_ws_sheet_row_heights(
    workspace_id: str,
    sheet_id: str,
    body: RowHeightsUpdate,
    request: Request,
    current_user: User = Depends(require_admin),
    db: DBSession = Depends(get_db),
):
    """워크스페이스 시트 행 높이 저장"""
    ws_sheet = db.query(WorkspaceSheet).filter(
        WorkspaceSheet.id == sheet_id,
        WorkspaceSheet.workspace_id == workspace_id,
    ).first()
    if not ws_sheet:
        raise HTTPException(status_code=404, detail="Sheet not found")
    ws_sheet.row_heights = json.dumps(body.row_heights) if body.row_heights else None
    db.commit()
    import asyncio
    asyncio.create_task(hub.broadcast(workspace_id, {
        "type": "sheet_config_updated",
        "sheet_id": sheet_id,
        "updated_by": current_user.username,
        "tab_id": request.headers.get("X-Tab-ID", ""),
    }, exclude=None))
    return {"message": "row_heights saved"}


class ColWidthsUpdate(BaseModel):
    col_widths: Optional[dict] = None  # e.g. {"0": 150, "3": 200} (col_index_str → px)


@admin_router.patch("/{workspace_id}/sheets/{sheet_id}/col-widths")
async def update_ws_sheet_col_widths(
    workspace_id: str,
    sheet_id: str,
    body: ColWidthsUpdate,
    request: Request,
    current_user: User = Depends(require_admin),
    db: DBSession = Depends(get_db),
):
    """워크스페이스 시트 열 너비 저장"""
    ws_sheet = db.query(WorkspaceSheet).filter(
        WorkspaceSheet.id == sheet_id,
        WorkspaceSheet.workspace_id == workspace_id,
    ).first()
    if not ws_sheet:
        raise HTTPException(status_code=404, detail="Sheet not found")
    ws_sheet.col_widths = json.dumps(body.col_widths) if body.col_widths else None
    db.commit()
    import asyncio
    asyncio.create_task(hub.broadcast(workspace_id, {
        "type": "sheet_config_updated",
        "sheet_id": sheet_id,
        "updated_by": current_user.username,
        "tab_id": request.headers.get("X-Tab-ID", ""),
    }, exclude=None))
    return {"message": "col_widths saved"}


@admin_router.patch("/{workspace_id}/sheets/{sheet_id}/freeze")
async def update_ws_sheet_freeze(
    workspace_id: str,
    sheet_id: str,
    body: FreezeUpdate,
    request: Request,
    current_user: User = Depends(require_admin),
    db: DBSession = Depends(get_db),
):
    """워크스페이스 시트 틀 고정 설정"""
    ws_sheet = db.query(WorkspaceSheet).filter(
        WorkspaceSheet.id == sheet_id,
        WorkspaceSheet.workspace_id == workspace_id,
    ).first()
    if not ws_sheet:
        raise HTTPException(status_code=404, detail="Sheet not found")
    ws_sheet.freeze_panes = body.freeze_panes
    db.commit()
    import asyncio
    asyncio.create_task(hub.broadcast(workspace_id, {
        "type": "sheet_config_updated",
        "sheet_id": sheet_id,
        "updated_by": current_user.username,
        "tab_id": request.headers.get("X-Tab-ID", ""),
    }, exclude=None))
    return {"message": "freeze saved", "freeze_panes": body.freeze_panes}


@admin_router.patch("/{workspace_id}/sheets/{sheet_id}/conditional-formats")
async def update_ws_conditional_formats(
    workspace_id: str,
    sheet_id: str,
    body: ConditionalFormatsUpdate,
    request: Request,
    current_user: User = Depends(require_admin),
    db: DBSession = Depends(get_db),
):
    """워크스페이스 시트 조건부 서식 저장"""
    ws_sheet = db.query(WorkspaceSheet).filter(
        WorkspaceSheet.id == sheet_id,
        WorkspaceSheet.workspace_id == workspace_id,
    ).first()
    if not ws_sheet:
        raise HTTPException(status_code=404, detail="Sheet not found")
    ws_sheet.conditional_formats = body.conditional_formats
    db.commit()
    import asyncio
    asyncio.create_task(hub.broadcast(workspace_id, {
        "type": "sheet_config_updated",
        "sheet_id": sheet_id,
        "updated_by": current_user.username,
        "tab_id": request.headers.get("X-Tab-ID", ""),
    }, exclude=None))
    return {"message": "conditional formats saved"}


@admin_router.post("/{workspace_id}/sheets", status_code=201)
async def add_workspace_sheet(
    workspace_id: str,
    body: WSSheetCreate,
    current_user: User = Depends(require_admin),
    db: DBSession = Depends(get_db),
):
    ws = db.query(Workspace).filter(Workspace.id == workspace_id).first()
    if not ws:
        raise HTTPException(status_code=404, detail="Workspace not found")
    if len(ws.sheets) >= 64:
        raise HTTPException(status_code=400, detail="Max 64 sheets")

    # 템플릿이 삭제된 경우 새 템플릿 생성
    tmpl = db.query(Template).filter(Template.id == ws.template_id).first() if ws.template_id else None
    if not tmpl:
        tmpl = Template(
            id=str(uuid.uuid4()), name=f"_auto_{ws.name}",
            created_by=current_user.id, updated_at=_now(),
        )
        db.add(tmpl)
        db.flush()
        ws.template_id = tmpl.id

    next_tmpl_idx = max((s.sheet_index for s in tmpl.sheets), default=-1) + 1
    new_tmpl_sheet = TemplateSheet(
        id=str(uuid.uuid4()), template_id=tmpl.id,
        sheet_index=next_tmpl_idx, sheet_name=body.sheet_name,
    )
    db.add(new_tmpl_sheet)
    for i, h in enumerate(["A", "B", "C", "D", "E"]):
        db.add(TemplateColumn(id=str(uuid.uuid4()), sheet_id=new_tmpl_sheet.id, col_index=i, col_header=h))

    next_ws_idx = max((s.sheet_index for s in ws.sheets), default=-1) + 1
    ws_sheet = WorkspaceSheet(
        id=str(uuid.uuid4()), workspace_id=workspace_id,
        template_sheet_id=new_tmpl_sheet.id,
        sheet_index=next_ws_idx, sheet_name=body.sheet_name,
    )
    db.add(ws_sheet)
    ws.updated_at = _now()
    db.commit()
    db.refresh(ws_sheet)

    import asyncio
    sheet_info = {
        "id": ws_sheet.id,
        "sheet_index": ws_sheet.sheet_index,
        "sheet_name": ws_sheet.sheet_name,
        "template_sheet_id": ws_sheet.template_sheet_id,
        "columns": [{"id": str(uuid.uuid4()), "col_index": i, "col_header": h,
                     "col_type": "text", "is_readonly": 0, "width": 120}
                    for i, h in enumerate(["A", "B", "C", "D", "E"])],
    }
    asyncio.create_task(hub.broadcast(workspace_id, {"type": "sheet_added", "sheet": sheet_info}))
    return {"data": sheet_info, "message": "created"}


@admin_router.patch("/{workspace_id}/sheets/{sheet_id}")
async def update_workspace_sheet(
    workspace_id: str,
    sheet_id: str,
    body: WSSheetUpdate,
    current_user: User = Depends(require_admin),
    db: DBSession = Depends(get_db),
):
    ws_sheet = db.query(WorkspaceSheet).filter(
        WorkspaceSheet.id == sheet_id,
        WorkspaceSheet.workspace_id == workspace_id,
    ).first()
    if not ws_sheet:
        raise HTTPException(status_code=404, detail="Sheet not found")
    if body.sheet_name is not None:
        ws_sheet.sheet_name = body.sheet_name
    db.commit()
    import asyncio
    asyncio.create_task(hub.broadcast(workspace_id, {
        "type": "sheet_renamed", "sheet_id": sheet_id, "sheet_name": ws_sheet.sheet_name,
    }))
    return {"data": {"id": ws_sheet.id, "sheet_name": ws_sheet.sheet_name}, "message": "updated"}


@admin_router.delete("/{workspace_id}/sheets/{sheet_id}", status_code=204)
async def delete_workspace_sheet(
    workspace_id: str,
    sheet_id: str,
    current_user: User = Depends(require_admin),
    db: DBSession = Depends(get_db),
):
    ws = db.query(Workspace).filter(Workspace.id == workspace_id).first()
    if not ws:
        raise HTTPException(status_code=404, detail="Workspace not found")
    if len(ws.sheets) <= 1:
        raise HTTPException(status_code=400, detail="Cannot delete the last sheet")
    ws_sheet = db.query(WorkspaceSheet).filter(
        WorkspaceSheet.id == sheet_id,
        WorkspaceSheet.workspace_id == workspace_id,
    ).first()
    if not ws_sheet:
        raise HTTPException(status_code=404, detail="Sheet not found")
    db.delete(ws_sheet)
    ws.updated_at = _now()
    db.commit()
    import asyncio
    asyncio.create_task(hub.broadcast(workspace_id, {"type": "sheet_deleted", "sheet_id": sheet_id}))


@admin_router.post("/{workspace_id}/import-xlsx")
async def import_workspace_xlsx(
    workspace_id: str,
    file: UploadFile = File(...),
    current_user: User = Depends(require_admin),
    db: DBSession = Depends(get_db),
):
    ws = db.query(Workspace).filter(Workspace.id == workspace_id).first()
    if not ws:
        raise HTTPException(status_code=404, detail="Workspace not found")

    content = await file.read()
    if content[:4] != b"PK\x03\x04":
        raise HTTPException(status_code=400, detail="Invalid xlsx file")

    content = _sanitize_xlsx(content)
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=False)
    theme_colors = _get_theme_colors(wb)
    ws_sheets = sorted(ws.sheets, key=lambda s: s.sheet_index)

    tmpl_sheet_ids = [s.template_sheet_id for s in ws_sheets if s.template_sheet_id]
    tmpl_sheets_map: dict[str, TemplateSheet] = {}
    if tmpl_sheet_ids:
        for ts in db.query(TemplateSheet).filter(TemplateSheet.id.in_(tmpl_sheet_ids)).all():
            tmpl_sheets_map[ts.id] = ts

    for ws_sheet in ws_sheets:
        if ws_sheet.sheet_name not in wb.sheetnames:
            continue
        excel_ws = wb[ws_sheet.sheet_name]

        # 기존 셀 삭제
        db.query(WorkspaceCell).filter(WorkspaceCell.sheet_id == ws_sheet.id).delete()

        all_rows = list(excel_ws.iter_rows(values_only=False))

        tmpl_sheet = tmpl_sheets_map.get(ws_sheet.template_sheet_id)
        num_cols = len(tmpl_sheet.columns) if tmpl_sheet else 0
        # xlsx 실제 컬럼 수도 반영 (열 삽입으로 확장된 경우 포함)
        num_cols = max(num_cols, excel_ws.max_column or 0, 1)

        # 병합, 행높이, 틀고정 업데이트
        merges_list = [str(mr) for mr in excel_ws.merged_cells.ranges]
        ws_sheet.merges = json.dumps(merges_list) if merges_list else None

        row_heights: dict = {}
        for row_num, rd in excel_ws.row_dimensions.items():
            if rd.height and rd.height > 0:
                row_heights[str(row_num - 1)] = rd.height
        ws_sheet.row_heights = json.dumps(row_heights) if row_heights else None

        col_widths: dict = {}
        for ci in range(num_cols):
            col_letter = get_column_letter(ci + 1)
            col_dim = excel_ws.column_dimensions.get(col_letter)
            if col_dim and col_dim.width and col_dim.width > 0:
                col_widths[str(ci)] = max(40, min(int(col_dim.width * 7), 600))
        ws_sheet.col_widths = json.dumps(col_widths) if col_widths else None

        ws_sheet.freeze_panes = str(excel_ws.freeze_panes) if excel_ws.freeze_panes else None

        for ri, row in enumerate(all_rows[:10000]):
            for ci in range(num_cols):
                cell = row[ci] if ci < len(row) else None
                if cell is None or cell.value is None:
                    style_json = _extract_cell_style(cell, theme_colors) if cell else None
                    comment_text = cell.comment.text.strip() if cell and cell.comment else None
                    if style_json or comment_text:
                        db.add(WorkspaceCell(
                            id=str(uuid.uuid4()), sheet_id=ws_sheet.id,
                            row_index=ri, col_index=ci,
                            value=None, style=style_json, comment=comment_text,
                            updated_by=current_user.id, updated_at=_now(),
                        ))
                    continue
                raw = cell.value
                style_json = _extract_cell_style(cell, theme_colors)
                comment_text = cell.comment.text.strip() if cell.comment else None
                if cell.data_type == 'f' or (isinstance(raw, str) and raw.startswith("=")):
                    raw_str = str(raw)
                    val = raw_str if raw_str.startswith("=") else "=" + raw_str
                else:
                    val = _stringify_value(raw)
                db.add(WorkspaceCell(
                    id=str(uuid.uuid4()),
                    sheet_id=ws_sheet.id,
                    row_index=ri,
                    col_index=ci,
                    value=val,
                    style=style_json,
                    comment=comment_text,
                    updated_by=current_user.id,
                    updated_at=_now(),
                ))

    db.commit()
    import asyncio
    asyncio.create_task(hub.broadcast(workspace_id, {"type": "reload"}))
    return {"message": "imported"}


@admin_router.get("/{workspace_id}/export-xlsx")
async def export_workspace_xlsx(
    workspace_id: str,
    current_user: User = Depends(require_admin),
    db: DBSession = Depends(get_db),
):
    ws = db.query(Workspace).filter(Workspace.id == workspace_id).first()
    if not ws:
        raise HTTPException(status_code=404, detail="Workspace not found")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for ws_sheet in sorted(ws.sheets, key=lambda s: s.sheet_index):
        excel_ws = wb.create_sheet(ws_sheet.sheet_name)
        tmpl_sheet = db.query(TemplateSheet).filter(TemplateSheet.id == ws_sheet.template_sheet_id).first()
        cols = sorted(tmpl_sheet.columns, key=lambda c: c.col_index) if tmpl_sheet else []

        # 컬럼 너비 (col_widths 우선, 없으면 TemplateColumn.width 사용)
        cw_override = {}
        if ws_sheet.col_widths:
            try:
                cw_override = json.loads(ws_sheet.col_widths)
            except Exception:
                pass
        for col in cols:
            w = int(cw_override.get(str(col.col_index), col.width or 120))
            excel_ws.column_dimensions[get_column_letter(col.col_index + 1)].width = w / 7

        # 행 높이 복원 (pt)
        if ws_sheet.row_heights:
            try:
                for ri_str, height_pt in json.loads(ws_sheet.row_heights).items():
                    excel_ws.row_dimensions[int(ri_str) + 1].height = float(height_pt)
            except Exception:
                pass

        # 틀 고정
        if ws_sheet.freeze_panes:
            excel_ws.freeze_panes = ws_sheet.freeze_panes

        # 셀 데이터 + 스타일 + 메모
        cells = db.query(WorkspaceCell).filter(WorkspaceCell.sheet_id == ws_sheet.id).all()
        for c in cells:
            val = c.value
            if val is not None and not val.startswith("="):
                val = _parse_value_for_excel(val)
            excel_cell = excel_ws.cell(row=c.row_index + 1, column=c.col_index + 1, value=val)
            _apply_cell_style(excel_cell, c.style)
            if c.comment:
                from openpyxl.comments import Comment as XlComment
                excel_cell.comment = XlComment(c.comment, "")

        # 병합 셀 복원
        if ws_sheet.merges:
            try:
                for rng in json.loads(ws_sheet.merges):
                    excel_ws.merge_cells(rng)
            except Exception:
                pass

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    safe_name = ws.name.replace(" ", "_")[:50]
    from urllib.parse import quote
    encoded_name = quote(safe_name + ".xlsx")
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_name}"},
    )


# ---------------------------------------------------------------
# User routes
# ---------------------------------------------------------------

user_router = APIRouter(prefix="/api/workspaces")


@user_router.get("")
async def list_workspaces(
    current_user: User = Depends(require_user),
    db: DBSession = Depends(get_db),
):
    workspaces = db.query(Workspace).order_by(Workspace.created_at.desc()).all()
    return {"data": [_ws_summary(w) for w in workspaces]}


@user_router.get("/{workspace_id}")
async def get_workspace(
    workspace_id: str,
    current_user: User = Depends(require_user),
    db: DBSession = Depends(get_db),
):
    ws = db.query(Workspace).filter(Workspace.id == workspace_id).first()
    if not ws:
        raise HTTPException(status_code=404, detail="Workspace not found")
    return {"data": _ws_detail(ws, db)}


router.include_router(admin_router)
router.include_router(user_router)
