# SPDX-License-Identifier: MIT
# Copyright (c) 2026 JAEHYUK CHO
import io
import json
import re
import uuid
import zipfile
import xml.etree.ElementTree as _ET
from datetime import datetime as _datetime, date as _date, time as _time, timedelta as _td
from typing import Optional
from urllib.parse import quote as url_quote
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy.orm import Session as DBSession
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string, range_boundaries
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection

from ..database import get_db
from ..models import (
    Template, TemplateSheet, TemplateColumn, TemplateCell, _now
)
from ..rbac import require_admin
from ..models import User

router = APIRouter(prefix="/api/admin/templates", tags=["templates"])

MAX_SHEETS = 64
MAX_ROWS = 10000


# ── Pydantic models ──────────────────────────────────────────
class TemplateCreate(BaseModel):
    name: str
    description: Optional[str] = None


class TemplateUpdate(BaseModel):
    name: Optional[str] = None
    description: Optional[str] = None


class SheetCreate(BaseModel):
    sheet_name: str = "새 시트"


class SheetUpdate(BaseModel):
    sheet_name: Optional[str] = None
    sheet_index: Optional[int] = None


class ColumnCreate(BaseModel):
    col_header: str = "새 컬럼"
    col_type: str = "text"
    is_readonly: int = 0
    width: int = 120


class ColumnUpdate(BaseModel):
    col_header: Optional[str] = None
    col_type: Optional[str] = None
    col_options: Optional[str] = None
    is_readonly: Optional[int] = None
    width: Optional[int] = None


class CellBatchItem(BaseModel):
    row_index: int
    col_index: int
    value: Optional[str] = None
    style: Optional[str] = None
    comment: Optional[str] = None


class MergesUpdate(BaseModel):
    merges: list[str]   # list of xlsx range strings e.g. ["A1:B2", "C3:D4"]


# ── xlsx sanitiser ───────────────────────────────────────────
_OXML_NS = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"

def _sanitize_xlsx(raw: bytes) -> bytes:
    """규격 위반 xlsx 수정 (font family > 14 등). 수정 필요 없으면 원본 반환."""
    try:
        with zipfile.ZipFile(io.BytesIO(raw)) as zin:
            if "xl/styles.xml" not in zin.namelist():
                return raw
            styles_xml = zin.read("xl/styles.xml")
    except zipfile.BadZipFile:
        return raw

    root = _ET.fromstring(styles_xml)
    fonts_el = root.find(f"{_OXML_NS}fonts")
    need_fix = False
    if fonts_el is not None:
        for font_el in fonts_el:
            fam = font_el.find(f"{_OXML_NS}family")
            if fam is not None:
                v = int(fam.get("val", "0"))
                if v > 14:
                    fam.set("val", "2")  # 2 = Swiss (sans-serif)
                    need_fix = True

    if not need_fix:
        return raw

    new_styles = _ET.tostring(root, xml_declaration=True, encoding="UTF-8")
    buf = io.BytesIO()
    with zipfile.ZipFile(io.BytesIO(raw)) as zin, zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.namelist():
            zout.writestr(item, new_styles if item == "xl/styles.xml" else zin.read(item))
    return buf.getvalue()


# ── Theme & color helpers ────────────────────────────────────
def _get_theme_colors(wb) -> list[str]:
    """워크북 테마 색상 팔레트 추출. theme index → 6자리 RGB hex (대문자)."""
    # 기본 Office 테마 (lt1, dk1, lt2, dk2, accent1-6, hlink, folHlink)
    defaults = ['FFFFFF', '000000', 'E7E6E6', '44546A',
                '4472C4', 'ED7D31', 'A5A5A5', 'FFC000',
                '5B9BD5', '70AD47', '0563C1', '954F72']
    try:
        theme_xml = wb.loaded_theme
        if not theme_xml:
            return defaults
        ns = {'a': 'http://schemas.openxmlformats.org/drawingml/2006/main'}
        root = _ET.fromstring(theme_xml)
        scheme = root.find('.//a:themeElements/a:clrScheme', ns)
        if scheme is None:
            return defaults
        # XML 순서: dk1, lt1, dk2, lt2, accent1-6, hlink, folHlink
        tag_order = ['dk1', 'lt1', 'dk2', 'lt2',
                     'accent1', 'accent2', 'accent3', 'accent4',
                     'accent5', 'accent6', 'hlink', 'folHlink']
        parsed: list[str] = []
        for tag in tag_order:
            el = scheme.find(f'a:{tag}', ns)
            if el is not None:
                srgb = el.find('a:srgbClr', ns)
                sys_clr = el.find('a:sysClr', ns)
                if srgb is not None:
                    parsed.append(srgb.get('val', '000000').upper())
                elif sys_clr is not None:
                    parsed.append(sys_clr.get('lastClr', '000000').upper())
                else:
                    parsed.append(defaults[len(parsed)] if len(parsed) < len(defaults) else '000000')
            else:
                parsed.append(defaults[len(parsed)] if len(parsed) < len(defaults) else '000000')
        # Excel theme index ↔ clrScheme 순서 스왑: (dk1,lt1,dk2,lt2) → (lt1,dk1,lt2,dk2)
        if len(parsed) >= 4:
            parsed[0], parsed[1] = parsed[1], parsed[0]
            parsed[2], parsed[3] = parsed[3], parsed[2]
        return parsed
    except Exception:
        return defaults


def _apply_tint(rgb_hex: str, tint: float) -> str:
    """Excel tint 적용. tint: -1.0(어둡게) ~ 1.0(밝게)."""
    if abs(tint) < 0.001:
        return rgb_hex.upper()
    r, g, b = int(rgb_hex[0:2], 16), int(rgb_hex[2:4], 16), int(rgb_hex[4:6], 16)
    if tint < 0:
        factor = 1.0 + tint
        r, g, b = int(r * factor), int(g * factor), int(b * factor)
    else:
        r = int(r + (255 - r) * tint)
        g = int(g + (255 - g) * tint)
        b = int(b + (255 - b) * tint)
    return f"{min(255, max(0, r)):02X}{min(255, max(0, g)):02X}{min(255, max(0, b)):02X}"


def _resolve_color(color_obj, theme_colors: list[str] | None = None) -> str | None:
    """openpyxl Color → 6자리 RGB hex (대문자). 실패 시 None."""
    if color_obj is None:
        return None
    try:
        if color_obj.type == 'rgb':
            rgb = color_obj.rgb or ''
            if len(rgb) >= 6:
                return rgb[-6:].upper()
        elif color_obj.type == 'theme' and theme_colors:
            idx = color_obj.theme
            tint = color_obj.tint or 0.0
            if idx is not None and 0 <= idx < len(theme_colors):
                return _apply_tint(theme_colors[idx], tint)
        elif color_obj.type == 'indexed':
            idx = color_obj.indexed
            if idx is not None:
                try:
                    from openpyxl.styles.colors import COLOR_INDEX
                    if 0 <= idx < len(COLOR_INDEX):
                        argb = COLOR_INDEX[idx]
                        if len(argb) >= 6:
                            return argb[-6:].upper()
                except Exception:
                    pass
    except Exception:
        pass
    return None


_ISO_DT_RE = re.compile(r'^\d{4}-\d{2}-\d{2}[T ]\d{2}:\d{2}:\d{2}$')
_ISO_DATE_RE = re.compile(r'^\d{4}-\d{2}-\d{2}$')


def _parse_value_for_excel(val_str: str | None):
    """DB 저장 문자열 → Excel에 적합한 Python 타입 변환."""
    if val_str is None:
        return None
    if val_str.startswith("="):
        return val_str
    # 불리언
    if val_str == 'TRUE':
        return True
    if val_str == 'FALSE':
        return False
    # 선행 0 보존 (텍스트 유지)
    if len(val_str) > 1 and val_str[0] == '0' and val_str != '0' and not val_str.startswith('0.'):
        return val_str
    # 정수
    try:
        if '.' not in val_str and 'e' not in val_str.lower():
            return int(val_str)
    except (ValueError, TypeError):
        pass
    # 실수
    try:
        return float(val_str)
    except (ValueError, TypeError):
        pass
    # datetime
    if _ISO_DT_RE.match(val_str):
        try:
            return _datetime.fromisoformat(val_str.replace(' ', 'T'))
        except ValueError:
            pass
    elif _ISO_DATE_RE.match(val_str):
        try:
            return _date.fromisoformat(val_str)
        except ValueError:
            pass
    return val_str


def _stringify_value(raw) -> str:
    """openpyxl 셀 값 → 저장용 문자열."""
    if isinstance(raw, bool):
        return 'TRUE' if raw else 'FALSE'
    if isinstance(raw, _datetime):
        return raw.isoformat(sep=' ')
    if isinstance(raw, _date):
        return raw.isoformat()
    if isinstance(raw, _time):
        return raw.isoformat()
    if isinstance(raw, _td):
        total = int(raw.total_seconds())
        h, rem = divmod(abs(total), 3600)
        m, s = divmod(rem, 60)
        sign = '-' if total < 0 else ''
        return f"{sign}{h:02d}:{m:02d}:{s:02d}"
    return str(raw)


# ── Style helpers ─────────────────────────────────────────────
def _extract_cell_style(cell, theme_colors=None) -> Optional[str]:
    """openpyxl cell → style JSON 문자열 (변경된 속성만)"""
    style: dict = {}

    # Font
    font = cell.font
    if font:
        if font.name and font.name != 'Calibri':
            style['fontName'] = font.name
        if font.bold:
            style['bold'] = True
        if font.italic:
            style['italic'] = True
        if font.underline and font.underline != 'none':
            # 밑줄 타입 보존: single, double, singleAccounting, doubleAccounting
            style['underline'] = font.underline if isinstance(font.underline, str) else 'single'
        if font.strikethrough:
            style['strikethrough'] = True
        if font.size and font.size != 11:
            style['fontSize'] = font.size
        # 폰트 색상: theme/indexed/rgb 모두 해석
        try:
            fc = _resolve_color(font.color, theme_colors)
            if fc and fc != '000000':
                style['color'] = fc
        except Exception:
            pass

    # Fill
    fill = cell.fill
    if fill and getattr(fill, 'fill_type', None) and fill.fill_type != 'none':
        try:
            fg = _resolve_color(fill.fgColor, theme_colors)
            if fg and fg != 'FFFFFF':
                # 검정(000000) 배경도 유효 → 제외하지 않음
                style['bg'] = fg
        except Exception:
            pass

    # Alignment
    align = cell.alignment
    if align:
        if align.horizontal and align.horizontal not in ('general', None):
            style['align'] = align.horizontal
        if align.vertical and align.vertical not in ('bottom', None):
            style['valign'] = align.vertical
        if align.wrap_text:
            style['wrap'] = True
        if align.indent and align.indent > 0:
            style['indent'] = align.indent
        if align.text_rotation and align.text_rotation != 0:
            style['textRotation'] = align.text_rotation

    # Border
    border = cell.border
    if border:
        borders: dict = {}
        for side_name in ('top', 'bottom', 'left', 'right'):
            side = getattr(border, side_name, None)
            if side and side.border_style and side.border_style != 'none':
                entry: dict = {'style': side.border_style}
                try:
                    bc = _resolve_color(side.color, theme_colors)
                    entry['color'] = bc if bc else '000000'
                except Exception:
                    entry['color'] = '000000'
                borders[side_name] = entry
        if borders:
            style['border'] = borders

    # Number format
    num_fmt = cell.number_format
    if num_fmt and num_fmt != 'General':
        style['numFmt'] = num_fmt

    # Cell protection (locked)
    prot = cell.protection
    if prot and prot.locked:
        style['locked'] = True

    return json.dumps(style, ensure_ascii=False) if style else None


def _apply_cell_style(ws_cell, style_json: Optional[str]) -> None:
    """style JSON 문자열 → openpyxl cell 스타일 적용"""
    if not style_json:
        return
    try:
        style = json.loads(style_json)
    except Exception:
        return
    if not style:
        return

    def _to_argb(hex6: str) -> str:
        """6자리 RGB hex → 8자리 ARGB hex (불투명)"""
        return ('FF' + hex6) if len(hex6) == 6 else hex6

    # Font
    font_kwargs: dict = {}
    if style.get('fontName'):
        font_kwargs['name'] = style['fontName']
    if style.get('bold'):
        font_kwargs['bold'] = True
    if style.get('italic'):
        font_kwargs['italic'] = True
    ul = style.get('underline')
    if ul:
        # 하위호환: bool(True) 또는 문자열('single','double',...)
        font_kwargs['underline'] = ul if isinstance(ul, str) else 'single'
    if style.get('strikethrough'):
        font_kwargs['strike'] = True
    if style.get('fontSize'):
        font_kwargs['size'] = style['fontSize']
    if style.get('color'):
        font_kwargs['color'] = _to_argb(style['color'])
    if font_kwargs:
        ws_cell.font = Font(**font_kwargs)

    # Fill
    if style.get('bg'):
        ws_cell.fill = PatternFill(fill_type='solid', fgColor=_to_argb(style['bg']))

    # Alignment
    align_kwargs: dict = {}
    if style.get('align'):
        align_kwargs['horizontal'] = style['align']
    if style.get('valign'):
        va = style['valign']
        if va == 'middle':
            va = 'center'
        align_kwargs['vertical'] = va
    if style.get('wrap'):
        align_kwargs['wrap_text'] = True
    if style.get('indent'):
        align_kwargs['indent'] = style['indent']
    if style.get('textRotation'):
        align_kwargs['text_rotation'] = style['textRotation']
    if align_kwargs:
        ws_cell.alignment = Alignment(**align_kwargs)

    # Border
    if style.get('border'):
        sides: dict = {}
        for side_name, bd in style['border'].items():
            color = _to_argb(bd.get('color', '000000'))
            sides[side_name] = Side(border_style=bd.get('style', 'thin'), color=color)
        ws_cell.border = Border(**sides)

    # Number format
    if style.get('numFmt'):
        ws_cell.number_format = style['numFmt']

    # Cell protection
    if style.get('locked'):
        from openpyxl.styles import Protection
        ws_cell.protection = Protection(locked=True)


def _style_to_css(style: dict) -> str:
    """style dict → CSS 문자열"""
    parts = []
    if style.get('fontName'):
        parts.append(f"font-family:{style['fontName']}")
    if style.get('bold'):
        parts.append('font-weight:bold')
    if style.get('italic'):
        parts.append('font-style:italic')
    decorations = []
    if style.get('underline'):
        decorations.append('underline')
    if style.get('strikethrough'):
        decorations.append('line-through')
    if decorations:
        parts.append('text-decoration:' + ' '.join(decorations))
    if style.get('fontSize'):
        parts.append(f"font-size:{style['fontSize']}pt")
    if style.get('color'):
        parts.append(f"color:#{style['color']}")
    if style.get('bg'):
        parts.append(f"background-color:#{style['bg']}")
    if style.get('align'):
        parts.append(f"text-align:{style['align']}")
    if style.get('valign'):
        va = style['valign']
        if va == 'center':
            va = 'middle'
        parts.append(f"vertical-align:{va}")
    if style.get('wrap'):
        parts.append('white-space:pre-wrap')
    if style.get('indent'):
        px = int(style['indent']) * 8
        parts.append(f"padding-left:{px}px")
    if style.get('textRotation'):
        parts.append(f"--text-rotation:{style['textRotation']}")
    if style.get('border'):
        width_map = {'thin': '1px', 'medium': '2px', 'thick': '3px',
                     'hair': '1px', 'dashed': '1px', 'dotted': '1px', 'double': '3px',
                     'mediumDashed': '2px', 'dashDot': '1px', 'mediumDashDot': '2px',
                     'dashDotDot': '1px', 'mediumDashDotDot': '2px', 'slantDashDot': '2px'}
        style_map = {'thin': 'solid', 'medium': 'solid', 'thick': 'solid',
                     'hair': 'solid', 'dashed': 'dashed', 'dotted': 'dotted', 'double': 'double',
                     'mediumDashed': 'dashed', 'dashDot': 'dashed', 'mediumDashDot': 'dashed',
                     'dashDotDot': 'dashed', 'mediumDashDotDot': 'dashed', 'slantDashDot': 'dashed'}
        for side, bd in style['border'].items():
            bs = bd.get('style', 'thin')
            color = bd.get('color', '000000')
            w = width_map.get(bs, '1px')
            s = style_map.get(bs, 'solid')
            parts.append(f"border-{side}:{w} {s} #{color}")
    if style.get('numFmt'):
        parts.append(f"--num-fmt:{url_quote(style['numFmt'], safe='')}")
    return ';'.join(parts)


def _range_to_jss(range_str: str) -> Optional[tuple]:
    """xlsx 병합 범위 "A1:B2" → (cell_name, [colspan, rowspan])"""
    try:
        min_col, min_row, max_col, max_row = range_boundaries(range_str)
        cell_name = f"{get_column_letter(min_col)}{min_row}"
        colspan = max_col - min_col + 1
        rowspan = max_row - min_row + 1
        return cell_name, [colspan, rowspan]
    except Exception:
        return None


def _freeze_to_cols(freeze_str: Optional[str]) -> int:
    """xlsx freeze_panes 문자열 → 고정 열 수"""
    if not freeze_str:
        return 0
    m = re.match(r'^([A-Z]+)\d+$', freeze_str.upper())
    if m:
        return column_index_from_string(m.group(1)) - 1
    return 0


def _pt_to_px(pt: float) -> int:
    """포인트 → 픽셀 (96dpi 기준)"""
    return round(pt * 96 / 72)


def _freeze_to_rows(freeze_str: Optional[str]) -> int:
    """xlsx freeze_panes 문자열 → 고정 행 수"""
    if not freeze_str:
        return 0
    m = re.match(r'^[A-Z]+(\d+)$', freeze_str.upper())
    if not m:
        return 0
    return max(0, int(m.group(1)) - 1)


def _extract_conditional_formats(worksheet) -> list:
    """openpyxl 워크시트에서 조건부 서식 규칙 추출 → JSON 배열"""
    rules = []
    try:
        for cf in worksheet.conditional_formatting:
            range_str = str(cf)
            for rule in cf.rules:
                entry = {"range": range_str, "type": rule.type}
                if rule.type == "cellIs":
                    entry["operator"] = rule.operator
                    if rule.formula:
                        entry["formula"] = [str(f) for f in rule.formula]
                elif rule.type == "expression":
                    if rule.formula:
                        entry["formula"] = [str(f) for f in rule.formula]
                elif rule.type == "colorScale":
                    entry["type"] = "colorScale"
                    try:
                        cs = rule.colorScale
                        colors = []
                        for c in (cs.color or []):
                            rgb = str(c.rgb) if c.rgb else "000000"
                            if len(rgb) == 8: rgb = rgb[2:]
                            colors.append(rgb)
                        entry["colors"] = colors
                    except Exception:
                        entry["colors"] = []
                    rules.append(entry)
                    continue
                elif rule.type == "dataBar":
                    entry["type"] = "dataBar"
                    try:
                        db_rule = rule.dataBar
                        color_rgb = "4472C4"
                        if db_rule and db_rule.color and db_rule.color.rgb:
                            c = str(db_rule.color.rgb)
                            color_rgb = c[2:] if len(c) == 8 else c
                        entry["color"] = color_rgb
                    except Exception:
                        entry["color"] = "4472C4"
                    rules.append(entry)
                    continue
                elif rule.type == "iconSet":
                    entry["type"] = "iconSet"
                    try:
                        entry["iconStyle"] = rule.iconSet.iconSet if rule.iconSet else "3TrafficLights1"
                    except Exception:
                        entry["iconStyle"] = "3TrafficLights1"
                    rules.append(entry)
                    continue
                else:
                    continue  # 지원하지 않는 타입
                # 서식 추출
                if rule.dxf:
                    fmt = {}
                    if rule.dxf.font:
                        if rule.dxf.font.bold:
                            fmt["bold"] = True
                        if rule.dxf.font.italic:
                            fmt["italic"] = True
                        if rule.dxf.font.color and rule.dxf.font.color.rgb:
                            rgb = str(rule.dxf.font.color.rgb)
                            if len(rgb) == 8:
                                rgb = rgb[2:]
                            fmt["color"] = rgb
                    if rule.dxf.fill and rule.dxf.fill.fgColor and rule.dxf.fill.fgColor.rgb:
                        rgb = str(rule.dxf.fill.fgColor.rgb)
                        if len(rgb) == 8:
                            rgb = rgb[2:]
                        fmt["bg"] = rgb
                    entry["format"] = fmt
                rules.append(entry)
    except Exception:
        pass
    return rules


def _apply_conditional_formats(worksheet, rules: list):
    """JSON 규칙 배열을 openpyxl 워크시트에 적용"""
    from openpyxl.formatting.rule import CellIsRule, FormulaRule
    from openpyxl.styles import Font, PatternFill
    for rule in rules:
        try:
            range_str = rule.get("range", "A1:A1")
            fmt = rule.get("format", {})
            font = None
            fill = None
            if fmt.get("bold") or fmt.get("italic") or fmt.get("color"):
                font = Font(
                    bold=fmt.get("bold", False),
                    italic=fmt.get("italic", False),
                    color=("FF" + fmt["color"]) if fmt.get("color") else None,
                )
            if fmt.get("bg"):
                fill = PatternFill(start_color="FF" + fmt["bg"], end_color="FF" + fmt["bg"], fill_type="solid")

            if rule.get("type") == "cellIs":
                formula_vals = rule.get("formula", [])
                worksheet.conditional_formatting.add(
                    range_str,
                    CellIsRule(
                        operator=rule.get("operator", "equal"),
                        formula=formula_vals,
                        font=font, fill=fill,
                    )
                )
            elif rule.get("type") == "expression":
                formula_vals = rule.get("formula", [])
                if formula_vals:
                    worksheet.conditional_formatting.add(
                        range_str,
                        FormulaRule(
                            formula=[formula_vals[0]],
                            font=font, fill=fill,
                        )
                    )
            elif rule.get("type") == "colorScale":
                from openpyxl.formatting.rule import ColorScaleRule
                colors_hex = rule.get("colors", [])
                if len(colors_hex) == 2:
                    worksheet.conditional_formatting.add(
                        range_str,
                        ColorScaleRule(
                            start_color="FF" + colors_hex[0],
                            end_color="FF" + colors_hex[1],
                        )
                    )
                elif len(colors_hex) >= 3:
                    worksheet.conditional_formatting.add(
                        range_str,
                        ColorScaleRule(
                            start_color="FF" + colors_hex[0],
                            mid_color="FF" + colors_hex[1],
                            end_color="FF" + colors_hex[2],
                        )
                    )
            elif rule.get("type") == "dataBar":
                from openpyxl.formatting.rule import DataBarRule
                bar_color = rule.get("color", "4472C4")
                worksheet.conditional_formatting.add(
                    range_str,
                    DataBarRule(
                        start_type="min", end_type="max",
                        color="FF" + bar_color,
                    )
                )
            elif rule.get("type") == "iconSet":
                from openpyxl.formatting.rule import IconSetRule
                icon_style = rule.get("iconStyle", "3TrafficLights1")
                worksheet.conditional_formatting.add(
                    range_str,
                    IconSetRule(
                        icon_style=icon_style,
                        type="percent",
                        values=[0, 33, 67],
                    )
                )
        except Exception:
            continue


def _extract_data_validations(worksheet) -> list:
    """openpyxl 워크시트에서 데이터 유효성 규칙 추출"""
    rules = []
    try:
        for dv in worksheet.data_validations.dataValidation:
            rule = {
                "ranges": str(dv.sqref),
                "type": dv.type or "none",
                "operator": dv.operator,
                "allowBlank": dv.allow_blank,
                "showDropDown": dv.showDropDown,
                "errorTitle": dv.errorTitle,
                "error": dv.error,
                "errorStyle": dv.errorStyle,
                "promptTitle": dv.promptTitle,
                "prompt": dv.prompt,
            }
            if dv.formula1:
                rule["formula1"] = str(dv.formula1)
            if dv.formula2:
                rule["formula2"] = str(dv.formula2)
            rules.append(rule)
    except Exception:
        pass
    return rules


def _apply_data_validations(worksheet, rules: list):
    """JSON 규칙을 openpyxl 데이터 유효성으로 적용"""
    from openpyxl.worksheet.datavalidation import DataValidation
    for rule in rules:
        try:
            dv = DataValidation(
                type=rule.get("type", "none"),
                operator=rule.get("operator"),
                formula1=rule.get("formula1"),
                formula2=rule.get("formula2"),
                allow_blank=rule.get("allowBlank", True),
                showDropDown=rule.get("showDropDown"),
                errorTitle=rule.get("errorTitle"),
                error=rule.get("error"),
                errorStyle=rule.get("errorStyle", "stop"),
                promptTitle=rule.get("promptTitle"),
                prompt=rule.get("prompt"),
            )
            dv.sqref = rule.get("ranges", "A1")
            worksheet.add_data_validation(dv)
        except Exception:
            continue


def _extract_outlines(worksheet) -> tuple:
    """openpyxl 워크시트에서 행/열 그룹(outline) 레벨 추출"""
    row_outlines = {}
    col_outlines = {}
    try:
        for ri, rd in worksheet.row_dimensions.items():
            if rd.outline_level and rd.outline_level > 0:
                row_outlines[str(ri - 1)] = rd.outline_level  # 0-based
    except Exception:
        pass
    try:
        for col_letter, cd in worksheet.column_dimensions.items():
            if cd.outline_level and cd.outline_level > 0:
                ci = column_index_from_string(col_letter) - 1
                col_outlines[str(ci)] = cd.outline_level
    except Exception:
        pass
    return (row_outlines, col_outlines)


def _apply_outlines(worksheet, row_outlines: dict, col_outlines: dict):
    """JSON outline → openpyxl 행/열 그룹 적용"""
    if row_outlines:
        for ri_str, level in row_outlines.items():
            try:
                ri = int(ri_str) + 1  # 1-based
                worksheet.row_dimensions[ri].outline_level = level
            except Exception:
                continue
    if col_outlines:
        for ci_str, level in col_outlines.items():
            try:
                ci = int(ci_str)
                col_letter = get_column_letter(ci + 1)
                worksheet.column_dimensions[col_letter].outline_level = level
            except Exception:
                continue


# ── Helpers ──────────────────────────────────────────────────
def _template_summary(t: Template) -> dict:
    return {
        "id": t.id,
        "name": t.name,
        "description": t.description,
        "created_by": t.created_by,
        "created_at": t.created_at,
        "updated_at": t.updated_at,
        "sheet_count": len(t.sheets),
    }


def _col_dict(c: TemplateColumn) -> dict:
    return {
        "id": c.id,
        "col_index": c.col_index,
        "col_header": c.col_header,
        "col_type": c.col_type,
        "col_options": json.loads(c.col_options) if c.col_options else None,
        "is_readonly": c.is_readonly,
        "width": c.width,
    }


def _sheet_detail(sheet: TemplateSheet) -> dict:
    return {
        "id": sheet.id,
        "sheet_index": sheet.sheet_index,
        "sheet_name": sheet.sheet_name,
        "columns": [_col_dict(c) for c in sheet.columns],
    }


def _default_columns(sheet_id: str):
    return [
        TemplateColumn(
            id=str(uuid.uuid4()), sheet_id=sheet_id,
            col_index=i, col_header=h,
        )
        for i, h in enumerate(["A", "B", "C", "D", "E"])
    ]


# ── Template CRUD ─────────────────────────────────────────────
@router.get("")
async def list_templates(
    current_user: User = Depends(require_admin),
    db: DBSession = Depends(get_db),
):
    templates = db.query(Template).order_by(Template.created_at.desc()).all()
    return {"data": [_template_summary(t) for t in templates]}


@router.post("", status_code=201)
async def create_template(
    body: TemplateCreate,
    current_user: User = Depends(require_admin),
    db: DBSession = Depends(get_db),
):
    t = Template(id=str(uuid.uuid4()), name=body.name, description=body.description,
                 created_by=current_user.id)
    db.add(t)
    sheet = TemplateSheet(id=str(uuid.uuid4()), template_id=t.id, sheet_index=0, sheet_name="Sheet1")
    db.add(sheet)
    for col in _default_columns(sheet.id):
        db.add(col)
    db.commit()
    db.refresh(t)
    return {"data": _template_summary(t), "message": "created"}


@router.get("/{template_id}")
async def get_template(
    template_id: str,
    current_user: User = Depends(require_admin),
    db: DBSession = Depends(get_db),
):
    t = db.query(Template).filter(Template.id == template_id).first()
    if not t:
        raise HTTPException(status_code=404, detail="Template not found")
    return {"data": {**_template_summary(t), "sheets": [_sheet_detail(s) for s in t.sheets]}}


@router.patch("/{template_id}")
async def update_template(
    template_id: str,
    body: TemplateUpdate,
    current_user: User = Depends(require_admin),
    db: DBSession = Depends(get_db),
):
    t = db.query(Template).filter(Template.id == template_id).first()
    if not t:
        raise HTTPException(status_code=404, detail="Template not found")
    if body.name is not None:
        t.name = body.name
    if body.description is not None:
        t.description = body.description
    t.updated_at = _now()
    db.commit()
    return {"data": _template_summary(t), "message": "updated"}


@router.delete("/{template_id}", status_code=204)
async def delete_template(
    template_id: str,
    current_user: User = Depends(require_admin),
    db: DBSession = Depends(get_db),
):
    t = db.query(Template).filter(Template.id == template_id).first()
    if not t:
        raise HTTPException(status_code=404, detail="Template not found")
    db.delete(t)
    db.commit()


class BatchDeleteRequest(BaseModel):
    ids: list[str]


@router.post("/batch-delete")
async def batch_delete_templates(
    body: BatchDeleteRequest,
    current_user: User = Depends(require_admin),
    db: DBSession = Depends(get_db),
):
    targets = db.query(Template).filter(Template.id.in_(body.ids)).all()
    for t in targets:
        db.delete(t)
    db.commit()
    return {"deleted": len(targets)}


@router.post("/{template_id}/copy", status_code=201)
async def copy_template(
    template_id: str,
    current_user: User = Depends(require_admin),
    db: DBSession = Depends(get_db),
):
    src = db.query(Template).filter(Template.id == template_id).first()
    if not src:
        raise HTTPException(status_code=404, detail="Template not found")
    new_t = Template(id=str(uuid.uuid4()), name=f"{src.name} (copy)",
                     description=src.description, created_by=current_user.id)
    db.add(new_t)
    for sheet in src.sheets:
        new_sheet = TemplateSheet(
            id=str(uuid.uuid4()), template_id=new_t.id,
            sheet_index=sheet.sheet_index, sheet_name=sheet.sheet_name,
            merges=sheet.merges, row_heights=sheet.row_heights,
            col_widths=sheet.col_widths, freeze_panes=sheet.freeze_panes,
            conditional_formats=sheet.conditional_formats,
        )
        db.add(new_sheet)
        for col in sheet.columns:
            db.add(TemplateColumn(id=str(uuid.uuid4()), sheet_id=new_sheet.id,
                                  col_index=col.col_index, col_header=col.col_header,
                                  col_type=col.col_type, col_options=col.col_options,
                                  is_readonly=col.is_readonly, width=col.width))
        for cell in sheet.cells:
            db.add(TemplateCell(id=str(uuid.uuid4()), sheet_id=new_sheet.id,
                                row_index=cell.row_index, col_index=cell.col_index,
                                value=cell.value, formula=cell.formula, style=cell.style,
                                comment=cell.comment))
    db.commit()
    db.refresh(new_t)
    return {"data": _template_summary(new_t), "message": "copied"}


# ── Sheet CRUD ────────────────────────────────────────────────
@router.get("/{template_id}/sheets/{sheet_id}/snapshot")
async def get_template_sheet_snapshot(
    template_id: str,
    sheet_id: str,
    current_user: User = Depends(require_admin),
    db: DBSession = Depends(get_db),
):
    """서식 편집기용 셀 스냅샷 반환. 병합·스타일·행높이·틀고정 포함."""
    sheet = db.query(TemplateSheet).filter(
        TemplateSheet.id == sheet_id,
        TemplateSheet.template_id == template_id,
    ).first()
    if not sheet:
        raise HTTPException(status_code=404, detail="Sheet not found")

    cols = sorted(sheet.columns, key=lambda c: c.col_index)
    num_cols = len(cols)
    if num_cols == 0:
        return {"data": {"cells": [], "num_rows": 50, "num_cols": 0,
                         "merges": {}, "row_heights": {}, "freeze_columns": 0, "styles": {}}}

    cells = sheet.cells
    max_row = max((c.row_index for c in cells), default=-1) + 1
    num_rows = max(max_row, 50)

    grid = [[""] * num_cols for _ in range(num_rows)]
    styles: dict = {}
    num_formats: dict = {}
    for c in cells:
        if c.row_index < num_rows and c.col_index < num_cols:
            val = c.formula if c.formula else (c.value or "")
            grid[c.row_index][c.col_index] = val
            if c.style:
                try:
                    s = json.loads(c.style)
                    if s:
                        cell_name = f"{get_column_letter(c.col_index + 1)}{c.row_index + 1}"
                        styles[cell_name] = _style_to_css(s)
                        if s.get('numFmt'):
                            num_formats[cell_name] = s['numFmt']
                except Exception:
                    pass

    # 병합 셀: xlsx range strings → jspreadsheet format
    merges: dict = {}
    if sheet.merges:
        try:
            for rng in json.loads(sheet.merges):
                result = _range_to_jss(rng)
                if result:
                    cell_name, dims = result
                    merges[cell_name] = dims
        except Exception:
            pass

    # 행 높이: pt → px
    row_heights_px: dict = {}
    if sheet.row_heights:
        try:
            for k, v in json.loads(sheet.row_heights).items():
                row_heights_px[k] = _pt_to_px(float(v))
        except Exception:
            pass

    # 열 너비: col_widths (px)
    col_widths_px: dict = {}
    if sheet.col_widths:
        try:
            col_widths_px = json.loads(sheet.col_widths)
        except Exception:
            pass

    # 틀 고정
    freeze_columns = _freeze_to_cols(sheet.freeze_panes)
    freeze_rows = _freeze_to_rows(sheet.freeze_panes)

    # 셀 메모
    comments: dict = {}
    for c in cells:
        if c.comment and c.row_index < num_rows and c.col_index < num_cols:
            cell_name = f"{get_column_letter(c.col_index + 1)}{c.row_index + 1}"
            comments[cell_name] = c.comment

    # 조건부 서식
    conditional_formats = []
    if sheet.conditional_formats:
        try:
            conditional_formats = json.loads(sheet.conditional_formats)
        except Exception:
            pass

    # 데이터 유효성 검사
    data_validations = []
    if sheet.data_validations:
        try:
            data_validations = json.loads(sheet.data_validations)
        except Exception:
            pass

    return {
        "data": {
            "cells": grid,
            "num_rows": num_rows,
            "num_cols": num_cols,
            "merges": merges,
            "row_heights": row_heights_px,
            "col_widths": col_widths_px,
            "freeze_columns": freeze_columns,
            "freeze_rows": freeze_rows,
            "styles": styles,
            "comments": comments,
            "num_formats": num_formats,
            "conditional_formats": conditional_formats,
            "data_validations": data_validations,
        }
    }


@router.post("/{template_id}/sheets", status_code=201)
async def add_template_sheet(
    template_id: str,
    body: SheetCreate,
    current_user: User = Depends(require_admin),
    db: DBSession = Depends(get_db),
):
    t = db.query(Template).filter(Template.id == template_id).first()
    if not t:
        raise HTTPException(status_code=404, detail="Template not found")
    if len(t.sheets) >= MAX_SHEETS:
        raise HTTPException(status_code=400, detail=f"Max {MAX_SHEETS} sheets allowed")

    next_index = max((s.sheet_index for s in t.sheets), default=-1) + 1
    sheet = TemplateSheet(id=str(uuid.uuid4()), template_id=template_id,
                          sheet_index=next_index, sheet_name=body.sheet_name)
    db.add(sheet)
    for col in _default_columns(sheet.id):
        db.add(col)
    t.updated_at = _now()
    db.commit()
    db.refresh(sheet)
    return {"data": _sheet_detail(sheet), "message": "created"}


@router.patch("/{template_id}/sheets/{sheet_id}")
async def update_template_sheet(
    template_id: str,
    sheet_id: str,
    body: SheetUpdate,
    current_user: User = Depends(require_admin),
    db: DBSession = Depends(get_db),
):
    sheet = db.query(TemplateSheet).filter(
        TemplateSheet.id == sheet_id,
        TemplateSheet.template_id == template_id,
    ).first()
    if not sheet:
        raise HTTPException(status_code=404, detail="Sheet not found")
    if body.sheet_name is not None:
        sheet.sheet_name = body.sheet_name
    if body.sheet_index is not None:
        sheet.sheet_index = body.sheet_index
    db.commit()
    return {"data": _sheet_detail(sheet), "message": "updated"}


@router.patch("/{template_id}/sheets/{sheet_id}/merges")
async def update_sheet_merges(
    template_id: str,
    sheet_id: str,
    body: MergesUpdate,
    current_user: User = Depends(require_admin),
    db: DBSession = Depends(get_db),
):
    """병합 셀 목록 저장 (jspreadsheet onmerge 콜백에서 호출)"""
    sheet = db.query(TemplateSheet).filter(
        TemplateSheet.id == sheet_id,
        TemplateSheet.template_id == template_id,
    ).first()
    if not sheet:
        raise HTTPException(status_code=404, detail="Sheet not found")
    sheet.merges = json.dumps(body.merges, ensure_ascii=False)
    db.commit()
    return {"message": "merges saved", "count": len(body.merges)}


class RowHeightsUpdate(BaseModel):
    row_heights: Optional[dict] = None  # e.g. {"0": 30, "5": 50} (row_index_str → pt)


@router.patch("/{template_id}/sheets/{sheet_id}/row-heights")
async def update_sheet_row_heights(
    template_id: str,
    sheet_id: str,
    body: RowHeightsUpdate,
    current_user: User = Depends(require_admin),
    db: DBSession = Depends(get_db),
):
    """서식 시트 행 높이 저장"""
    sheet = db.query(TemplateSheet).filter(
        TemplateSheet.id == sheet_id,
        TemplateSheet.template_id == template_id,
    ).first()
    if not sheet:
        raise HTTPException(status_code=404, detail="Sheet not found")
    sheet.row_heights = json.dumps(body.row_heights) if body.row_heights else None
    db.commit()
    return {"message": "row_heights saved"}


class ColWidthsUpdate(BaseModel):
    col_widths: Optional[dict] = None  # e.g. {"0": 150, "3": 200} (col_index_str → px)


@router.patch("/{template_id}/sheets/{sheet_id}/col-widths")
async def update_sheet_col_widths(
    template_id: str,
    sheet_id: str,
    body: ColWidthsUpdate,
    current_user: User = Depends(require_admin),
    db: DBSession = Depends(get_db),
):
    """서식 시트 열 너비 저장"""
    sheet = db.query(TemplateSheet).filter(
        TemplateSheet.id == sheet_id,
        TemplateSheet.template_id == template_id,
    ).first()
    if not sheet:
        raise HTTPException(status_code=404, detail="Sheet not found")
    sheet.col_widths = json.dumps(body.col_widths) if body.col_widths else None
    db.commit()
    return {"message": "col_widths saved"}


class FreezeUpdate(BaseModel):
    freeze_panes: Optional[str] = None


@router.patch("/{template_id}/sheets/{sheet_id}/freeze")
async def update_sheet_freeze(
    template_id: str,
    sheet_id: str,
    body: FreezeUpdate,
    current_user: User = Depends(require_admin),
    db: DBSession = Depends(get_db),
):
    """서식 시트 틀 고정 설정"""
    sheet = db.query(TemplateSheet).filter(
        TemplateSheet.id == sheet_id,
        TemplateSheet.template_id == template_id,
    ).first()
    if not sheet:
        raise HTTPException(status_code=404, detail="Sheet not found")
    sheet.freeze_panes = body.freeze_panes
    db.commit()
    return {"message": "freeze saved", "freeze_panes": body.freeze_panes}


class ConditionalFormatsUpdate(BaseModel):
    conditional_formats: Optional[str] = None  # JSON string of rules array


@router.patch("/{template_id}/sheets/{sheet_id}/conditional-formats")
async def update_sheet_conditional_formats(
    template_id: str,
    sheet_id: str,
    body: ConditionalFormatsUpdate,
    current_user: User = Depends(require_admin),
    db: DBSession = Depends(get_db),
):
    sheet = db.query(TemplateSheet).filter(
        TemplateSheet.id == sheet_id,
        TemplateSheet.template_id == template_id,
    ).first()
    if not sheet:
        raise HTTPException(status_code=404, detail="Sheet not found")
    sheet.conditional_formats = body.conditional_formats
    db.commit()
    return {"message": "conditional formats saved"}


@router.delete("/{template_id}/sheets/{sheet_id}", status_code=204)
async def delete_template_sheet(
    template_id: str,
    sheet_id: str,
    current_user: User = Depends(require_admin),
    db: DBSession = Depends(get_db),
):
    t = db.query(Template).filter(Template.id == template_id).first()
    if not t:
        raise HTTPException(status_code=404, detail="Template not found")
    if len(t.sheets) <= 1:
        raise HTTPException(status_code=400, detail="Cannot delete the last sheet")
    sheet = db.query(TemplateSheet).filter(
        TemplateSheet.id == sheet_id,
        TemplateSheet.template_id == template_id,
    ).first()
    if not sheet:
        raise HTTPException(status_code=404, detail="Sheet not found")
    db.delete(sheet)
    t.updated_at = _now()
    db.commit()


# ── Column CRUD ───────────────────────────────────────────────
@router.post("/{template_id}/sheets/{sheet_id}/columns", status_code=201)
async def add_template_column(
    template_id: str,
    sheet_id: str,
    body: ColumnCreate,
    current_user: User = Depends(require_admin),
    db: DBSession = Depends(get_db),
):
    sheet = db.query(TemplateSheet).filter(
        TemplateSheet.id == sheet_id,
        TemplateSheet.template_id == template_id,
    ).first()
    if not sheet:
        raise HTTPException(status_code=404, detail="Sheet not found")
    next_idx = max((c.col_index for c in sheet.columns), default=-1) + 1
    if body.col_type not in ("text", "number", "date", "dropdown", "checkbox"):
        raise HTTPException(status_code=400, detail="Invalid col_type")
    col = TemplateColumn(
        id=str(uuid.uuid4()), sheet_id=sheet_id,
        col_index=next_idx, col_header=body.col_header,
        col_type=body.col_type, is_readonly=body.is_readonly, width=body.width,
    )
    db.add(col)
    db.commit()
    db.refresh(col)
    return {"data": _col_dict(col), "message": "created"}


@router.patch("/{template_id}/sheets/{sheet_id}/columns/{col_id}")
async def update_column(
    template_id: str,
    sheet_id: str,
    col_id: str,
    body: ColumnUpdate,
    current_user: User = Depends(require_admin),
    db: DBSession = Depends(get_db),
):
    col = db.query(TemplateColumn).filter(
        TemplateColumn.id == col_id,
        TemplateColumn.sheet_id == sheet_id,
    ).first()
    if not col:
        raise HTTPException(status_code=404, detail="Column not found")
    if body.col_header is not None:
        col.col_header = body.col_header
    if body.col_type is not None:
        col.col_type = body.col_type
    if body.col_options is not None:
        col.col_options = body.col_options
    if body.is_readonly is not None:
        col.is_readonly = body.is_readonly
    if body.width is not None:
        col.width = body.width
    db.commit()
    return {"data": _col_dict(col), "message": "updated"}


@router.delete("/{template_id}/sheets/{sheet_id}/columns/{col_id}", status_code=204)
async def delete_template_column(
    template_id: str,
    sheet_id: str,
    col_id: str,
    current_user: User = Depends(require_admin),
    db: DBSession = Depends(get_db),
):
    col = db.query(TemplateColumn).filter(
        TemplateColumn.id == col_id,
        TemplateColumn.sheet_id == sheet_id,
    ).first()
    if not col:
        raise HTTPException(status_code=404, detail="Column not found")
    deleted_idx = col.col_index
    db.delete(col)
    # 해당 컬럼 셀 삭제
    db.query(TemplateCell).filter(
        TemplateCell.sheet_id == sheet_id,
        TemplateCell.col_index == deleted_idx,
    ).delete()
    # 뒤 컬럼 인덱스 당기기
    remaining = db.query(TemplateColumn).filter(
        TemplateColumn.sheet_id == sheet_id,
        TemplateColumn.col_index > deleted_idx,
    ).all()
    for c in remaining:
        c.col_index -= 1
    # 뒤 셀 col_index 당기기
    later_cells = db.query(TemplateCell).filter(
        TemplateCell.sheet_id == sheet_id,
        TemplateCell.col_index > deleted_idx,
    ).all()
    for c in later_cells:
        c.col_index -= 1
    # 병합 범위 업데이트
    ts = db.query(TemplateSheet).filter(TemplateSheet.id == sheet_id).first()
    if ts and ts.merges:
        try:
            merges = json.loads(ts.merges)
            from .cells import _shift_merges_col
            updated = _shift_merges_col(merges, deleted_idx, -1)
            ts.merges = json.dumps(updated)
        except Exception:
            pass
    db.commit()


# ── Cell batch save ───────────────────────────────────────────
@router.post("/{template_id}/sheets/{sheet_id}/cells")
async def batch_save_cells(
    template_id: str,
    sheet_id: str,
    body: list[CellBatchItem],
    replace: bool = False,
    current_user: User = Depends(require_admin),
    db: DBSession = Depends(get_db),
):
    sheet = db.query(TemplateSheet).filter(
        TemplateSheet.id == sheet_id,
        TemplateSheet.template_id == template_id,
    ).first()
    if not sheet:
        raise HTTPException(status_code=404, detail="Sheet not found")

    # replace=true: 기존 셀 전부 삭제 후 새로 저장 (행/열 삽입·삭제 후 전체 동기화용)
    if replace:
        db.query(TemplateCell).filter(TemplateCell.sheet_id == sheet_id).delete()

    for item in body:
        if not (0 <= item.row_index < MAX_ROWS):
            continue
        # 계산식 여부 분리
        formula = None
        value = item.value
        if item.value and item.value.startswith("="):
            formula = item.value
            value = None

        existing = db.query(TemplateCell).filter(
            TemplateCell.sheet_id == sheet_id,
            TemplateCell.row_index == item.row_index,
            TemplateCell.col_index == item.col_index,
        ).first()
        if existing:
            existing.value = value
            existing.formula = formula
            if item.style is not None:
                existing.style = item.style
            if item.comment is not None:
                existing.comment = item.comment if item.comment else None
        else:
            db.add(TemplateCell(
                id=str(uuid.uuid4()),
                sheet_id=sheet_id,
                row_index=item.row_index,
                col_index=item.col_index,
                value=value,
                formula=formula,
                style=item.style,
                comment=item.comment if item.comment else None,
            ))
    t = db.query(Template).filter(Template.id == template_id).first()
    if t:
        t.updated_at = _now()
    db.commit()
    return {"message": "saved", "count": len(body)}


# ── xlsx import / export ──────────────────────────────────────
@router.post("/import-xlsx", status_code=201)
async def import_template_xlsx(
    file: UploadFile = File(...),
    name: Optional[str] = None,
    current_user: User = Depends(require_admin),
    db: DBSession = Depends(get_db),
):
    content = await file.read()
    if content[:4] != b"PK\x03\x04":
        raise HTTPException(status_code=400, detail="Invalid xlsx file")

    # data_only=False → 계산식 원본 보존
    content = _sanitize_xlsx(content)
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=False)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Cannot open xlsx: {e}")

    if len(wb.sheetnames) > MAX_SHEETS:
        raise HTTPException(status_code=400, detail=f"Too many sheets (max {MAX_SHEETS})")

    # 테마 색상 팔레트 추출 (theme 색상 해석용)
    theme_colors = _get_theme_colors(wb)

    template_name = name or (file.filename or "Imported").replace(".xlsx", "")
    t = Template(id=str(uuid.uuid4()), name=template_name, created_by=current_user.id)
    db.add(t)

    for sheet_idx, sheet_name in enumerate(wb.sheetnames):
        ws = wb[sheet_name]
        new_sheet = TemplateSheet(id=str(uuid.uuid4()), template_id=t.id,
                                  sheet_index=sheet_idx, sheet_name=sheet_name)
        db.add(new_sheet)

        all_rows = list(ws.iter_rows(values_only=False))
        if not all_rows:
            db.add(TemplateColumn(id=str(uuid.uuid4()), sheet_id=new_sheet.id, col_index=0, col_header="A"))
            continue

        first_row = all_rows[0]
        # xlsx 전체 시트의 최대 컬럼 수 사용 (첫 행만 보면 뒤쪽 행의 데이터 누락 가능)
        num_cols = ws.max_column or 0
        if num_cols == 0:
            num_cols = max((i for i, cell in enumerate(first_row) if cell.value is not None), default=-1) + 1
        if num_cols == 0:
            num_cols = len(first_row) if first_row else 5

        # 컬럼 헤더는 Excel 열 문자(A, B, C...)
        for ci in range(num_cols):
            col_letter = get_column_letter(ci + 1)
            # 실제 xlsx 열 너비 사용 (px 변환: 1 unit ≈ 7px)
            col_dim = ws.column_dimensions.get(col_letter)
            col_width = int((col_dim.width or 8) * 7) if col_dim else 56
            col_width = max(40, min(col_width, 600))
            db.add(TemplateColumn(
                id=str(uuid.uuid4()),
                sheet_id=new_sheet.id,
                col_index=ci,
                col_header=col_letter,
                width=col_width,
            ))

        # 병합 셀 저장
        merges_list = [str(mr) for mr in ws.merged_cells.ranges]
        if merges_list:
            new_sheet.merges = json.dumps(merges_list)

        # 행 높이 저장 (pt 단위)
        row_heights: dict = {}
        for row_num, rd in ws.row_dimensions.items():
            if rd.height and rd.height > 0:
                ri = row_num - 1  # row_index = Excel row - 1
                row_heights[str(ri)] = rd.height
        if row_heights:
            new_sheet.row_heights = json.dumps(row_heights)

        # 틀 고정
        if ws.freeze_panes:
            new_sheet.freeze_panes = str(ws.freeze_panes)

        # 숨겨진 행/열 추출
        hidden_rows = [row_num - 1 for row_num, rd in ws.row_dimensions.items()
                       if rd.hidden]
        if hidden_rows:
            new_sheet.hidden_rows = json.dumps(hidden_rows)
        hidden_cols = []
        for col_letter, cd in ws.column_dimensions.items():
            if cd.hidden:
                try:
                    from openpyxl.utils import column_index_from_string
                    hidden_cols.append(column_index_from_string(col_letter) - 1)
                except Exception:
                    pass
        if hidden_cols:
            new_sheet.hidden_cols = json.dumps(hidden_cols)

        # 조건부 서식 추출
        cf_rules = _extract_conditional_formats(ws)
        if cf_rules:
            new_sheet.conditional_formats = json.dumps(cf_rules)

        # 데이터 유효성 검사 추출
        dv_rules = _extract_data_validations(ws)
        if dv_rules:
            new_sheet.data_validations = json.dumps(dv_rules)

        # 행/열 그룹 (outline) 추출
        row_outlines, col_outlines = _extract_outlines(ws)
        if row_outlines:
            new_sheet.outline_rows = json.dumps(row_outlines)
        if col_outlines:
            new_sheet.outline_cols = json.dumps(col_outlines)

        # 모든 행을 데이터로 저장
        for ri, row in enumerate(all_rows[:MAX_ROWS]):
            for ci in range(num_cols):
                cell = row[ci] if ci < len(row) else None
                if cell is None or cell.value is None:
                    # 값 없어도 스타일이나 메모/하이퍼링크 있으면 저장
                    style_json = _extract_cell_style(cell, theme_colors) if cell else None
                    comment_text = cell.comment.text.strip() if cell and cell.comment else None
                    hyperlink_url = cell.hyperlink.target if cell and cell.hyperlink and cell.hyperlink.target else None
                    if style_json or comment_text or hyperlink_url:
                        db.add(TemplateCell(
                            id=str(uuid.uuid4()), sheet_id=new_sheet.id,
                            row_index=ri, col_index=ci,
                            value=None, style=style_json, comment=comment_text,
                            hyperlink=hyperlink_url,
                        ))
                    continue
                raw = cell.value
                style_json = _extract_cell_style(cell, theme_colors)
                comment_text = cell.comment.text.strip() if cell.comment else None
                hyperlink_url = cell.hyperlink.target if cell.hyperlink and cell.hyperlink.target else None
                if cell.data_type == 'f' or (isinstance(raw, str) and raw.startswith("=")):
                    raw_str = str(raw)
                    formula_str = raw_str if raw_str.startswith("=") else "=" + raw_str
                    db.add(TemplateCell(
                        id=str(uuid.uuid4()), sheet_id=new_sheet.id,
                        row_index=ri, col_index=ci,
                        value=None, formula=formula_str, style=style_json,
                        comment=comment_text, hyperlink=hyperlink_url,
                    ))
                else:
                    db.add(TemplateCell(
                        id=str(uuid.uuid4()), sheet_id=new_sheet.id,
                        row_index=ri, col_index=ci,
                        value=_stringify_value(raw), style=style_json,
                        comment=comment_text, hyperlink=hyperlink_url,
                    ))

    db.commit()
    db.refresh(t)
    return {"data": _template_summary(t), "message": "imported"}


@router.get("/{template_id}/export-xlsx")
async def export_template_xlsx(
    template_id: str,
    current_user: User = Depends(require_admin),
    db: DBSession = Depends(get_db),
):
    t = db.query(Template).filter(Template.id == template_id).first()
    if not t:
        raise HTTPException(status_code=404, detail="Template not found")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for sheet in t.sheets:
        ws = wb.create_sheet(sheet.sheet_name)
        cols = sorted(sheet.columns, key=lambda c: c.col_index)

        # 컬럼 너비 (col_widths 우선, 없으면 TemplateColumn.width 사용)
        cw_override = {}
        if sheet.col_widths:
            try:
                cw_override = json.loads(sheet.col_widths)
            except Exception:
                pass
        for col in cols:
            w = int(cw_override.get(str(col.col_index), col.width or 120))
            ws.column_dimensions[get_column_letter(col.col_index + 1)].width = w / 7

        # 행 높이 복원 (pt 단위)
        if sheet.row_heights:
            try:
                for ri_str, height_pt in json.loads(sheet.row_heights).items():
                    ws.row_dimensions[int(ri_str) + 1].height = float(height_pt)
            except Exception:
                pass

        # 틀 고정
        if sheet.freeze_panes:
            ws.freeze_panes = sheet.freeze_panes

        # 셀 데이터 + 스타일
        cells_map: dict[tuple, TemplateCell] = {(c.row_index, c.col_index): c for c in sheet.cells}
        max_row = max((c.row_index for c in sheet.cells), default=-1)
        for ri in range(max_row + 1):
            for ci in range(len(cols)):
                c = cells_map.get((ri, ci))
                if c:
                    val = c.formula if c.formula else c.value
                    if val is not None and not val.startswith("="):
                        val = _parse_value_for_excel(val)
                    ws_cell = ws.cell(row=ri + 1, column=ci + 1, value=val)
                    _apply_cell_style(ws_cell, c.style)
                    if c.comment:
                        from openpyxl.comments import Comment as XlComment
                        ws_cell.comment = XlComment(c.comment, "")
                    if c.hyperlink:
                        ws_cell.hyperlink = c.hyperlink

        # 숨겨진 행/열 복원
        if sheet.hidden_rows:
            try:
                for ri in json.loads(sheet.hidden_rows):
                    ws.row_dimensions[ri + 1].hidden = True
            except Exception:
                pass
        if sheet.hidden_cols:
            try:
                for ci in json.loads(sheet.hidden_cols):
                    ws.column_dimensions[get_column_letter(ci + 1)].hidden = True
            except Exception:
                pass

        # 병합 셀 복원
        if sheet.merges:
            try:
                for rng in json.loads(sheet.merges):
                    ws.merge_cells(rng)
            except Exception:
                pass

        # 조건부 서식 복원
        if sheet.conditional_formats:
            try:
                cf_rules = json.loads(sheet.conditional_formats)
                _apply_conditional_formats(ws, cf_rules)
            except Exception:
                pass

        # 데이터 유효성 검사 복원
        if sheet.data_validations:
            try:
                dv_rules = json.loads(sheet.data_validations)
                _apply_data_validations(ws, dv_rules)
            except Exception:
                pass

        # 행/열 그룹 복원
        row_ol = {}
        col_ol = {}
        if sheet.outline_rows:
            try: row_ol = json.loads(sheet.outline_rows)
            except: pass
        if sheet.outline_cols:
            try: col_ol = json.loads(sheet.outline_cols)
            except: pass
        if row_ol or col_ol:
            _apply_outlines(ws, row_ol, col_ol)

        # readonly 컬럼 셀 잠금 (시트 보호 활성화)
        readonly_col_indices = {col.col_index for col in cols if col.is_readonly}
        if readonly_col_indices:
            total_rows = ws.max_row or 1
            for r in range(1, total_rows + 1):
                for ci, col in enumerate(cols):
                    cell = ws.cell(row=r, column=ci + 1)
                    cell.protection = Protection(locked=(col.col_index in readonly_col_indices))
            ws.protection.sheet = True

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    safe_name = t.name.replace(" ", "_")[:50]
    from urllib.parse import quote
    encoded_name = quote(safe_name + ".xlsx")
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_name}"},
    )
